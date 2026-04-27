import pandas as pd
import numpy as np
import json
import os
import math
import calendar
from datetime import datetime, date
from collections import defaultdict

# =============================================================
#  Smart APS V13  —  Demand-Floor + Fixed-Machine Lockdown
#  =========================================================
#  CHANGES FROM V12:
#  -----------------
#  1. TERMINAL THRESHOLDS (category-aware):
#       Runner   → 95% × MIN_RUN_HOURS × rate  (pieces that can be made in MIN_RUN_HOURS)
#       Repeater → 85% × MIN_RUN_HOURS × rate
#       Stranger → 75% × MIN_RUN_HOURS × rate
#     Threshold is the minimum terminal inventory required to allow production.
#
#  2. FIXED MACHINES — ABSOLUTE LOCKDOWN:
#     Fixed machines ALWAYS run ONLY their fixed parts, every day, no exceptions.
#     - No Phase A / Phase B distinction (removed entirely).
#     - Inventory level of fixed parts is irrelevant — they always run.
#     - No OPD cap applied to fixed machine runs.
#     - Fixed machines are completely excluded from the general scheduler,
#       utilization enforcer, runner priority displacement, and strategic filler.
#
#  3. PRODUCTION QTY FLOOR:
#     Target for every part = max(indent_daily, demand_daily).
#     effective_daily already captures this; now enforced as a hard floor
#     in assign_part — the scheduler will not stop early just because
#     inventory already covers indent if demand is higher.
#
#  4. PRIORITY SCORING — HIGH DEMAND / HIGH INDENT FIRST:
#     Parts with high effective_daily AND low inventory vs safety stock
#     are ranked highest. New urgency formula:
#       urgency = 0.50 × gap_score + 0.30 × velocity + 0.20 × demand_weight
#     demand_weight = effective_daily / max_effective_daily (normalised)
#     This ensures high-volume parts with low coverage always go first.
# =============================================================


# =============================================================
# SECTION 1 — DEFAULT PARAMETERS
# =============================================================

DEFAULTS = dict(
    AVAILABLE_HOURS=22,
    AVAILABLE_HOURS_EXTENDED=23,
    MIN_RUN_HOURS=4,
    MIN_DAILY_INDENT=150,
    MIN_INDENT_HOURS=4.0,
    SAFETY_DAYS=3,
    TARGET_DAYS=5,
    OPD_SCENARIO_0=3.0,
    OPD_SCENARIO_1=3.0,
    OPD_SCENARIO_2=4.0,
    OPD_SCENARIO_3=5.0,
    W_URGENCY=0.55,
    W_CATEGORY=0.25,
    W_INDENT=0.20,
    UTIL_TARGET_PCT=90.0,
    COLOR_PURGE_HRS=10 / 60.0,
    RUNNER_PRIORITY_DAYS=2.0,
    MAX_DAILY_CO=25,
    FORWARD_LOOK_DAYS=7,
    # ── V13: category-specific terminal threshold multipliers ──
    TERMINAL_THRESHOLD={
        "Runner":   0.95,   # 95% of MIN_RUN_HOURS production capacity
        "Repeater": 0.85,   # 85%
        "Stranger": 0.75,   # 75%
    },
    STRATEGIC_BUFFER_DAYS=7,
    STRATEGIC_PRIORITY_DISCOUNT=0.5,
    EFFICIENCY_MODE_UTIL_FLOOR=95.0,
    ABSOLUTE_MAX_DAYS=15,
    SPECIALIZED_MACHINE_THRESHOLD=3,
)


# =============================================================
# SECTION 2 — WORKING DAYS HELPER
# =============================================================

def compute_working_days(ref_date):
    year = ref_date.year
    month = ref_date.month
    total = calendar.monthrange(year, month)[1]
    sundays = sum(
        1 for d in range(1, total + 1)
        if date(year, month, d).weekday() == 6
    )
    return total - sundays, total, sundays


# =============================================================
# SECTION 3 — COLUMN FINDER
# =============================================================

def find_col(df, name, sheet):
    match = next(
        (c for c in df.columns if str(c).strip().lower() == name.lower()), None
    )
    if match is None:
        raise ValueError(
            f"Column '{name}' not found in sheet '{sheet}'.\n"
            f"Available: {list(df.columns)}"
        )
    return match


def safe_dict(df, key_col, val_col, default=0.0):
    return {
        str(k).strip(): (default if pd.isna(v) else float(v))
        for k, v in zip(df[key_col], df[val_col])
        if pd.notna(k) and str(k).strip() != ""
    }


# =============================================================
# SECTION 4 — MAIN ENTRY POINT
# =============================================================

def run_aps(
    planning_date,
    indent_month,
    book_path,
    inventory_path,
    matrix_path,
    changeover_path,
    terminal_path,
    machine_state_path,
    output_path,
    section="VT",
    demand_path=None,
    params=None,
):
    """
    Run Smart APS V13.

    Parameters
    ----------
    planning_date       : datetime.date
    indent_month        : datetime.date  — first day of indent month
    book_path           : str/file-like — Book1 Excel (sheet "<section>")
    inventory_path      : str/file-like — Inventory Excel/CSV
    matrix_path         : str/file-like — compatibility matrix Excel
    changeover_path     : str/file-like — changeover times Excel
    terminal_path       : str/file-like/None — terminal data Excel
    machine_state_path  : str — path to machine_state.json (auto-created)
    output_path         : str — path to write output Excel
    section             : str — e.g. "VT" or "HZ"
    demand_path         : str/file-like/None
                          Excel with sheet "<section>_Demand" or CSV.
                          Columns: Part, Daily_Demand
    params              : dict/None — override any DEFAULTS key

    Returns
    -------
    dict: success, error, output_path, summary, sheets
    """

    section = (str(section).strip() if section is not None else "") or "VT"
    S = section

    P = dict(DEFAULTS)
    if params:
        P.update(params)

    AVAILABLE_HOURS          = P["AVAILABLE_HOURS"]
    AVAILABLE_HOURS_EXTENDED = P["AVAILABLE_HOURS_EXTENDED"]
    MIN_RUN_HOURS            = P["MIN_RUN_HOURS"]
    MIN_DAILY_INDENT         = P["MIN_DAILY_INDENT"]
    MIN_INDENT_HOURS         = P["MIN_INDENT_HOURS"]
    SAFETY_DAYS              = P["SAFETY_DAYS"]
    TARGET_DAYS              = P["TARGET_DAYS"]
    OPD_SCENARIO_0           = P["OPD_SCENARIO_0"]
    OPD_SCENARIO_1           = P["OPD_SCENARIO_1"]
    OPD_SCENARIO_2           = P["OPD_SCENARIO_2"]
    OPD_SCENARIO_3           = P["OPD_SCENARIO_3"]
    W_URGENCY                = P["W_URGENCY"]
    W_CATEGORY               = P["W_CATEGORY"]
    W_INDENT                 = P["W_INDENT"]
    UTIL_TARGET_PCT          = P["UTIL_TARGET_PCT"]
    COLOR_PURGE_HRS          = P["COLOR_PURGE_HRS"]
    RUNNER_PRIORITY_DAYS     = P["RUNNER_PRIORITY_DAYS"]
    MAX_DAILY_CO             = P["MAX_DAILY_CO"]
    FORWARD_LOOK_DAYS        = P["FORWARD_LOOK_DAYS"]
    TERMINAL_THRESHOLD       = P["TERMINAL_THRESHOLD"]   # now a dict {cat: pct}
    STRATEGIC_BUFFER_DAYS    = P["STRATEGIC_BUFFER_DAYS"]
    STRATEGIC_PRIORITY_DISCOUNT = P["STRATEGIC_PRIORITY_DISCOUNT"]
    EFFICIENCY_MODE_UTIL_FLOOR  = P["EFFICIENCY_MODE_UTIL_FLOOR"]
    ABSOLUTE_MAX_DAYS        = P["ABSOLUTE_MAX_DAYS"]
    SPECIALIZED_MACHINE_THRESHOLD = P["SPECIALIZED_MACHINE_THRESHOLD"]

    try:
        # =========================================================
        # LOAD DATA
        # =========================================================

        vt_parts_raw = pd.read_excel(book_path, sheet_name=S)

        if hasattr(inventory_path, "read") or (
            isinstance(inventory_path, str) and inventory_path.lower().endswith(".csv")
        ):
            inv_raw = pd.read_csv(inventory_path)
        else:
            inv_raw = pd.read_excel(inventory_path, sheet_name=f"{S}_Inventory")

        vt_matrix            = pd.read_excel(matrix_path, sheet_name=f"{S}_Matrix")
        vt_co_raw            = pd.read_excel(changeover_path, sheet_name=f"{S}_Changeover")
        vt_machine_count_raw = pd.read_excel(matrix_path, sheet_name=f"{S}_Machine_Part_Count")

        try:
            vt_fixed_raw = pd.read_excel(matrix_path, sheet_name=f"{S}_Fixed")
        except Exception:
            vt_fixed_raw = None

        vt_terminals_raw      = None
        vt_terminal_avail_raw = None
        if terminal_path is not None:
            try:
                vt_terminals_raw      = pd.read_excel(terminal_path, sheet_name=f"{S}_Terminals")
                vt_terminal_avail_raw = pd.read_excel(terminal_path, sheet_name=f"{S}_Terminal_Inventory")
            except Exception:
                vt_terminals_raw = None
                vt_terminal_avail_raw = None

        # ── Load demand file ───────────────────────────────────
        demand_raw_daily = {}
        if demand_path is not None:
            try:
                if hasattr(demand_path, "read") or (
                    isinstance(demand_path, str) and demand_path.lower().endswith(".csv")
                ):
                    dem_raw = pd.read_csv(demand_path)
                else:
                    dem_raw = pd.read_excel(demand_path, sheet_name=f"{S}_Demand")

                dem_col_part  = find_col(dem_raw, "Part",        f"{S}_Demand")
                dem_col_daily = find_col(dem_raw, "Daily_Demand", f"{S}_Demand")

                dem_data = dem_raw[
                    dem_raw[dem_col_part].notna() &
                    (dem_raw[dem_col_part].astype(str).str.strip() != "")
                ].copy()
                demand_raw_daily = {
                    str(k).strip(): (0.0 if pd.isna(v) else float(v))
                    for k, v in zip(dem_data[dem_col_part], dem_data[dem_col_daily])
                    if pd.notna(k) and str(k).strip() != ""
                }
            except Exception as dem_err:
                demand_raw_daily = {}

        # =========================================================
        # WORKING DAYS
        # =========================================================

        WORKING_DAYS, TOTAL_DAYS, SUNDAY_COUNT = compute_working_days(indent_month)

        # =========================================================
        # PARSE BOOK1
        # =========================================================

        vt_col_part      = find_col(vt_parts_raw, "Part",       S)
        vt_col_cycletime = find_col(vt_parts_raw, "Cycle time", S)
        vt_col_cavity    = find_col(vt_parts_raw, "Cavity",     S)
        vt_col_indent    = find_col(vt_parts_raw, "Indent",     S)
        vt_col_tools     = find_col(vt_parts_raw, "Tools",      S)
        vt_col_color     = find_col(vt_parts_raw, "Color",      S)

        try:
            vt_col_category = find_col(vt_parts_raw, "Category", S)
        except ValueError:
            vt_col_category = None

        data = vt_parts_raw[
            vt_parts_raw[vt_col_part].notna() &
            (vt_parts_raw[vt_col_part].astype(str).str.strip() != "")
        ].copy()
        data = data.drop_duplicates(subset=vt_col_part).copy()
        data["Material"] = data[vt_col_part].astype(str).str.strip()

        data["_ct"]  = pd.to_numeric(data[vt_col_cycletime], errors="coerce").replace(0, np.nan)
        data["Rate"] = 3600 / data["_ct"]

        data_valid     = data[data["Rate"].notna()].copy()
        data_zero_rate = data[data["Rate"].isna()].copy()

        # =========================================================
        # PARSE INVENTORY FILE
        # =========================================================

        inv_col_part = find_col(inv_raw, "Part",      f"{S}_Inventory")
        inv_col_inv  = find_col(inv_raw, "Inventory", f"{S}_Inventory")

        inv_data = inv_raw[
            inv_raw[inv_col_part].notna() &
            (inv_raw[inv_col_part].astype(str).str.strip() != "")
        ].copy()
        inv_data["Material"] = inv_data[inv_col_part].astype(str).str.strip()

        inventory_from_file = {
            str(k).strip(): (0.0 if pd.isna(v) else float(v))
            for k, v in zip(inv_data["Material"], inv_data[inv_col_inv])
            if pd.notna(k) and str(k).strip() != ""
        }

        # =========================================================
        # LOOKUP DICTIONARIES
        # =========================================================

        rate           = safe_dict(data_valid, "Material", "Rate")
        indent_monthly = safe_dict(data, "Material", vt_col_indent)

        tools_available = {}
        part_color      = {}

        for _, row in data.iterrows():
            p = str(row["Material"]).strip()
            v = row[vt_col_tools]
            tools_available[p] = (
                max(1, int(float(v)))
                if pd.notna(v) and str(v).strip() != ""
                else 1
            )
            c = row[vt_col_color]
            part_color[p] = (
                str(c).strip().upper()
                if pd.notna(c) and str(c).strip() not in ("", "nan")
                else "UNKNOWN"
            )

        inventory = {p: inventory_from_file.get(p, 0.0) for p in data["Material"]}
        for p, v in inventory_from_file.items():
            if p not in inventory:
                inventory[p] = v

        ALL_KNOWN_COLORS = dict(part_color)

        # ── indent_daily: policy-based daily target ────────────
        indent_daily = {
            p: round(qty / WORKING_DAYS, 4)
            for p, qty in indent_monthly.items()
        }

        # ── demand_daily per part (falls back to indent if absent) ──
        demand_daily = {
            p: demand_raw_daily.get(p, indent_daily.get(p, 0.0))
            for p in indent_daily
        }
        for p, v in demand_raw_daily.items():
            if p not in demand_daily:
                demand_daily[p] = v

        # ── V13: effective_daily = max(indent, demand) — hard floor ──
        # This is BOTH the production target and the minimum qty floor.
        # The scheduler must produce at least this much before stopping.
        effective_daily = {
            p: max(indent_daily.get(p, 0.0), demand_daily.get(p, 0.0))
            for p in set(list(indent_daily.keys()) + list(demand_daily.keys()))
        }

        today_target_qty = {
            p: max(0.0, effective_daily.get(p, 0.0) - inventory.get(p, 0.0))
            for p in effective_daily
        }

        # =========================================================
        # PART CATEGORY
        # =========================================================

        def build_category(df, cat_col_name):
            cat = {}
            part_col = next(
                (c for c in df.columns if str(c).strip().lower() == "part"), None
            )
            if part_col is None:
                return cat
            cat_col = cat_col_name
            for _, row in df.iterrows():
                part = row[part_col]
                if pd.isna(part) or str(part).strip() == "":
                    continue
                val = "Stranger"
                if cat_col and cat_col in df.columns and pd.notna(row[cat_col]):
                    val = str(row[cat_col]).strip().capitalize()
                    if val not in ("Runner", "Repeater", "Stranger"):
                        val = "Stranger"
                cat[str(part).strip()] = val
            return cat

        part_category  = build_category(vt_parts_raw, vt_col_category)
        CATEGORY_SCORE = {"Runner": 100, "Repeater": 60, "Stranger": 20}

        # =========================================================
        # FIXED MACHINE CONSTRAINT
        # =========================================================

        def build_fixed_machine_dicts(df):
            pfm, mfp = {}, {}
            if df is None or df.empty:
                return pfm, mfp
            machine_col = next(
                (c for c in df.columns if str(c).strip().lower() == "machine"), None
            )
            if machine_col is None:
                return pfm, mfp
            part_cols = [c for c in df.columns if str(c).strip().lower() != "machine"]
            if not part_cols:
                return pfm, mfp
            for _, row in df.iterrows():
                machine = row[machine_col]
                if pd.isna(machine) or str(machine).strip() == "":
                    continue
                m = str(machine).strip()
                for col in part_cols:
                    val = row[col]
                    if pd.isna(val) or str(val).strip() in ("", "nan"):
                        continue
                    p = str(val).strip()
                    if p in pfm:
                        continue
                    pfm[p] = m
                    mfp.setdefault(m, []).append(p)
            for m in mfp:
                mfp[m] = sorted(mfp[m])
            return pfm, mfp

        part_fixed_machine, machine_fixed_parts = build_fixed_machine_dicts(vt_fixed_raw)

        # =========================================================
        # TERMINAL CONSTRAINT  (V13 — category-aware thresholds)
        # =========================================================
        #
        # Threshold = category_pct × MIN_RUN_HOURS × rate_per_hour
        #
        # This means: the terminal must hold enough inventory to cover
        # the minimum viable run for that part, scaled by category priority.
        #   Runner   → 95% of a MIN_RUN_HOURS run worth of pieces
        #   Repeater → 85%
        #   Stranger → 75%
        # =========================================================

        def _build_part_terminals(df):
            result = {}
            if df is None or df.empty:
                return result
            part_col = next(
                (c for c in df.columns if str(c).strip().lower() in ("part", "material")), None
            )
            if part_col is None:
                return result
            terminal_cols = [
                c for c in df.columns
                if str(c).strip().lower() not in ("part", "material")
            ]
            for _, row in df.iterrows():
                part = row[part_col]
                if pd.isna(part) or str(part).strip() == "":
                    continue
                p = str(part).strip()
                terminals = []
                for col in terminal_cols:
                    val = row[col]
                    if pd.notna(val) and str(val).strip() not in ("", "nan"):
                        terminals.append(str(val).strip().upper())
                if terminals:
                    result[p] = sorted(terminals)
            return result

        def _build_terminal_status(df):
            result = {}
            if df is None or df.empty:
                return result
            term_col = next(
                (c for c in df.columns if str(c).strip().lower() == "terminal"), None
            )
            inv_col = next(
                (c for c in df.columns if str(c).strip().lower() == "inventory"), None
            )
            if term_col is None or inv_col is None:
                return result
            for _, row in df.iterrows():
                t = row[term_col]
                v = row[inv_col]
                if pd.isna(t) or str(t).strip() == "":
                    continue
                key = str(t).strip().upper()
                try:
                    qty = float(v) if pd.notna(v) else 0.0
                except (ValueError, TypeError):
                    qty = 0.0
                result[key] = qty
            return result

        part_terminals  = _build_part_terminals(vt_terminals_raw)
        terminal_status = _build_terminal_status(vt_terminal_avail_raw)

        def terminal_blocked(part, category_override=None):
            """
            V13: threshold = category_pct × MIN_RUN_HOURS × rate_per_hour
            Every category also has MIN_RUN_HOURS as an implicit lower bound
            on how much the terminal must support.
            """
            required = part_terminals.get(part, [])
            if not required:
                return False, ""

            cat       = category_override or part_category.get(part, "Stranger")
            r_val     = rate.get(part, 0.0)
            pct       = TERMINAL_THRESHOLD.get(cat, 0.75)  # default to Stranger

            # pieces the terminal must hold = pct × MIN_RUN_HOURS × rate
            # i.e. enough terminal stock to cover a minimum-length run
            threshold = pct * MIN_RUN_HOURS * r_val

            blocking = []
            for t in sorted(required):
                t_inv = terminal_status.get(t, 0)
                if t_inv < threshold:
                    blocking.append(
                        f"{t}(inv={t_inv:.0f} < need={threshold:.0f} "
                        f"[{cat} {int(pct*100)}% × {MIN_RUN_HOURS}h × {r_val:.1f}/h])"
                    )
            if blocking:
                return True, (
                    f"Terminal inadequate: {', '.join(blocking)}  "
                    f"(requires: {', '.join(sorted(required))})"
                )
            return False, ""

        # =========================================================
        # SKIP RULES  (use indent_daily — policy ceiling)
        # =========================================================

        def should_skip(part):
            i_daily  = indent_daily.get(part, 0.0)
            d_daily  = demand_daily.get(part, 0.0)
            eff      = effective_daily.get(part, 0.0)
            monthly  = indent_monthly.get(part, 0.0)
            r        = rate.get(part, 1.0)

            if i_daily <= MIN_DAILY_INDENT and d_daily <= MIN_DAILY_INDENT:
                return True, (
                    f"Effective daily {eff:.2f} <= {MIN_DAILY_INDENT} threshold "
                    f"(indent={i_daily:.2f}, demand={d_daily:.2f})"
                )

            indent_hrs = monthly / r if r > 0 else 0.0
            if indent_hrs <= MIN_INDENT_HOURS and d_daily <= MIN_DAILY_INDENT:
                return True, (
                    f"Monthly indent = {indent_hrs:.2f}h <= {MIN_INDENT_HOURS}h "
                    f"and demand {d_daily:.2f} <= threshold"
                )

            inv = inventory.get(part, 0.0)
            # V13: skip only when inventory covers TARGET_DAYS of effective demand
            # AND demand is also satisfied (not just indent)
            if eff > 0 and inv >= TARGET_DAYS * eff:
                return True, (
                    f"Inventory ({inv:.0f}) >= {TARGET_DAYS}-day effective target "
                    f"({TARGET_DAYS * eff:.0f} pcs) — at ceiling, skip today"
                )

            t_blocked, t_reason = terminal_blocked(part)
            if t_blocked:
                return True, t_reason

            return False, ""

        def is_hard_skip(part):
            return terminal_blocked(part)[0]

        def is_scheduling_skip(part):
            return should_skip(part)[0]

        # =========================================================
        # CHANGEOVER TIMES
        # =========================================================

        def build_changeover_dict(co_df):
            co_dict = {}
            for _, row in co_df.iterrows():
                machine = str(row["Unique Machines"]).strip()
                minutes = row["Changeover time"]
                if pd.notna(minutes) and machine:
                    co_dict[machine] = float(minutes) / 60.0
            return co_dict

        vt_changeover = build_changeover_dict(vt_co_raw)
        DEFAULT_CHANGEOVER_HRS = 40 / 60.0

        # =========================================================
        # MACHINE PART COUNT
        # =========================================================

        def build_machine_part_count(df):
            mpc = {}
            machine_col = next(
                (c for c in df.columns if str(c).strip().lower() == "machine"), None
            )
            count_col = next(
                (c for c in df.columns if str(c).strip().lower() == "part_count"), None
            )
            if machine_col is None or count_col is None:
                return {}
            for _, row in df.iterrows():
                m = str(row[machine_col]).strip()
                v = row[count_col]
                if m and pd.notna(v):
                    try:
                        mpc[m] = int(float(v))
                    except (ValueError, TypeError):
                        pass
            return mpc

        machine_part_count = build_machine_part_count(vt_machine_count_raw)
        max_part_count     = max(machine_part_count.values(), default=1) or 1

        # =========================================================
        # MACHINE STATE
        # =========================================================

        def load_machine_state():
            if os.path.exists(machine_state_path):
                try:
                    with open(machine_state_path) as f:
                        content = f.read().strip()
                    if not content:
                        os.remove(machine_state_path)
                        return {}
                    state = json.loads(content)
                    if not isinstance(state, dict):
                        os.remove(machine_state_path)
                        return {}
                    unknown = [p for p in state.values() if p not in part_color]
                    for p in unknown:
                        ALL_KNOWN_COLORS[p] = "NEEDS_PURGE"
                    return state
                except Exception:
                    os.remove(machine_state_path)
                    return {}
            return {}

        def save_machine_state(state):
            combined = {m: p for m, p in sorted(state.items()) if p is not None}
            with open(machine_state_path, "w") as f:
                json.dump(combined, f, indent=2, sort_keys=True)

        machine_state = load_machine_state()

        # =========================================================
        # COMPATIBILITY MATRIX
        # =========================================================

        def build_compatibility(matrix):
            machines = matrix.columns[1:].tolist()
            compat   = {}
            for _, row in matrix.iterrows():
                part = row["Part"]
                for m in machines:
                    if row[m] == 1:
                        compat.setdefault(part, []).append(m)
            for p in compat:
                compat[p] = sorted(compat[p])
            return compat, machines

        vt_compat, vt_machines = build_compatibility(vt_matrix)
        vt_machines = sorted(vt_machines)

        machine_compatible_parts = {m: [] for m in vt_machines}
        for p, machines in sorted(vt_compat.items()):
            for m in machines:
                if m in machine_compatible_parts:
                    machine_compatible_parts[m].append(p)
        for m in machine_compatible_parts:
            machine_compatible_parts[m] = sorted(machine_compatible_parts[m])

        # =========================================================
        # SCENARIO CLASSIFIER
        # =========================================================

        def classify_scenario(parts):
            coverage = []
            for p in sorted(parts):
                inv  = inventory.get(p, 0)
                eff  = effective_daily.get(p, 0)
                skip, _ = should_skip(p)
                if skip or eff == 0:
                    continue
                coverage.append(inv / eff)
            if not coverage:
                return 3, "SCENARIO 3 — No active parts today"
            n        = len(coverage)
            critical = sum(1 for c in coverage if c < 1)
            low      = sum(1 for c in coverage if c < SAFETY_DAYS)
            if critical == n:
                return 0, f"SCENARIO 0 — ALL {n} active parts critical (inv < 1 day)"
            elif critical > 0:
                return 1, f"SCENARIO 1 — {critical}/{n} parts critical  |  {n-critical} have buffer"
            elif low > 0:
                return 2, f"SCENARIO 2 — {low}/{n} parts below {SAFETY_DAYS}-day safety floor"
            else:
                return 3, f"SCENARIO 3 — All {n} parts healthy (>={SAFETY_DAYS} days safety floor)"

        def opd_cap(scenario_id):
            scenario_opd = {
                0: OPD_SCENARIO_0,
                1: OPD_SCENARIO_1,
                2: OPD_SCENARIO_2,
                3: OPD_SCENARIO_3,
            }.get(scenario_id, OPD_SCENARIO_2)
            return min(scenario_opd, TARGET_DAYS)

        def _opd_cap_qty(part, scenario_id, current_inventory):
            i_daily         = indent_daily.get(part, 0.0)
            eff             = effective_daily.get(part, 0.0)
            cap_from_indent = opd_cap(scenario_id) * i_daily
            # Never let the cap fall below one effective day of demand
            return max(cap_from_indent, eff)

        # =========================================================
        # PRIORITY SCORING  (V13 — high demand/high indent first)
        # =========================================================
        #
        # New urgency formula:
        #   urgency = 0.50 × gap_score + 0.30 × velocity_score + 0.20 × demand_weight
        #
        # demand_weight = effective_daily / max_effective_daily (normalised)
        # This ensures high-volume parts with low coverage rank highest.
        # Safety stock ratio (inv / SAFETY_DAYS*eff) acts as tie-breaker.
        # =========================================================

        def compute_priority_scores(active_parts):
            rows = []
            for p in sorted(active_parts):
                inv     = inventory.get(p, 0)
                eff     = effective_daily.get(p, 0)
                i_daily = indent_daily.get(p, 0)
                d_daily = demand_daily.get(p, 0)
                cat     = part_category.get(p, "Stranger")
                days_cov      = inv / eff if eff > 0 else 999.0
                gap_score     = min(1.0, max(0.0, (TARGET_DAYS - days_cov) / TARGET_DAYS))
                velocity_raw  = eff / max(float(inv), 1.0) if eff > 0 else 0.0
                # Safety stock pressure: how far below safety floor are we?
                safety_ratio  = (inv / (SAFETY_DAYS * eff)) if eff > 0 else 1.0
                rows.append({
                    "part": p, "inv": inv, "daily": eff,
                    "indent_daily": i_daily, "demand_daily": d_daily,
                    "days_cov": days_cov, "cat": cat,
                    "gap_score": gap_score, "velocity_raw": velocity_raw,
                    "safety_ratio": safety_ratio,
                })
            if not rows:
                return {}, []

            max_daily    = max(r["daily"] for r in rows) or 1.0
            max_velocity = max(r["velocity_raw"] for r in rows) or 1.0
            scores, score_rows = {}, []

            for r in rows:
                p = r["part"]

                gap_pct       = r["gap_score"] * 100.0
                velocity_pct  = (r["velocity_raw"] / max_velocity) * 100.0

                # V13: demand_weight rewards high effective_daily parts
                demand_weight = (r["daily"] / max_daily) * 100.0

                # V13 urgency: gap 50%, velocity 30%, demand volume 20%
                urgency_score = (
                    0.50 * gap_pct +
                    0.30 * velocity_pct +
                    0.20 * demand_weight
                )

                category_score = CATEGORY_SCORE.get(r["cat"], 20)

                # indent_score now uses effective_daily (max of indent+demand)
                indent_score = (r["daily"] / max_daily) * 100.0

                final_score = (
                    W_URGENCY  * urgency_score +
                    W_CATEGORY * category_score +
                    W_INDENT   * indent_score
                )

                # Safety pressure: parts below safety stock get a boost
                if r["safety_ratio"] < 1.0:
                    final_score *= (1.0 + (1.0 - r["safety_ratio"]) * 0.3)

                scores[p] = round(final_score, 2)
                score_rows.append({
                    "Part":              p,
                    "Category":          r["cat"],
                    "Color":             part_color.get(p, "UNKNOWN"),
                    "Fixed_Machine":     part_fixed_machine.get(p, "—"),
                    "Tools":             tools_available.get(p, 1),
                    "Inventory_Now":     round(r["inv"], 0),
                    "Indent_Daily":      round(r["indent_daily"], 2),
                    "Demand_Daily":      round(r["demand_daily"], 2),
                    "Effective_Daily":   round(r["daily"], 2),
                    "Days_Coverage":     round(r["days_cov"], 2),
                    "Safety_Ratio":      round(r["safety_ratio"], 3),
                    "Safety_Floor":      SAFETY_DAYS,
                    "Target_Ceiling":    TARGET_DAYS,
                    "Buffer_Status": (
                        "CRITICAL"     if r["days_cov"] < 1        else
                        "BELOW_SAFETY" if r["days_cov"] < SAFETY_DAYS else
                        "BUILDING"     if r["days_cov"] < TARGET_DAYS  else
                        "AT_TARGET"
                    ),
                    "Gap_Score_Pct":      round(gap_pct,      1),
                    "Velocity_Score_Pct": round(velocity_pct, 1),
                    "Demand_Weight_Pct":  round(demand_weight, 1),
                    "Urgency_Score":      round(urgency_score, 1),
                    "Category_Score":     category_score,
                    "Indent_Score":       round(indent_score,  1),
                    "Final_Score":        round(final_score,   2),
                })
            return scores, score_rows

        # =========================================================
        # COLOUR-AWARE CHANGEOVER HELPER
        # =========================================================

        def _co_hrs_for(part, machine, machine_last_part):
            last = machine_last_part.get(machine)
            if last is None or last == part:
                return 0.0
            base_co    = vt_changeover.get(machine, DEFAULT_CHANGEOVER_HRS)
            last_color = ALL_KNOWN_COLORS.get(last, "UNKNOWN")
            new_color  = part_color.get(part, "UNKNOWN")
            if last_color == "NEEDS_PURGE":
                return base_co + COLOR_PURGE_HRS
            purge = (
                COLOR_PURGE_HRS
                if last_color != new_color
                   and last_color not in ("UNKNOWN",)
                   and new_color  not in ("UNKNOWN",)
                else 0.0
            )
            return base_co + purge

        # =========================================================
        # MACHINE RANKER
        # =========================================================

        def _rank_normal(part, machines_to_try, machine_hours,
                         machine_last_part, new_color, runner_lock):
            ranked = []
            for m in sorted(machines_to_try):
                used = machine_hours.get(m, 0)
                free = round(AVAILABLE_HOURS - used, 4)
                if free < MIN_RUN_HOURS:
                    continue
                last = machine_last_part.get(m)
                if runner_lock and last != part:
                    continue
                if last is None or last == part:
                    co_hrs      = 0.0
                    color_bonus = 0.0
                else:
                    base_co    = vt_changeover.get(m, DEFAULT_CHANGEOVER_HRS)
                    last_color = ALL_KNOWN_COLORS.get(last, "UNKNOWN")
                    same_color = (
                        last_color == new_color
                        and last_color not in ("UNKNOWN", "NEEDS_PURGE")
                        and new_color  not in ("UNKNOWN",)
                    )
                    purge = (
                        0.0 if same_color else (
                            COLOR_PURGE_HRS
                            if last_color not in ("UNKNOWN",) and new_color not in ("UNKNOWN",)
                            else 0.0
                        )
                    )
                    co_hrs      = base_co + purge
                    color_bonus = -0.08 if same_color else 0.0
                effective_free = round(free - co_hrs, 4)
                if effective_free < MIN_RUN_HOURS:
                    continue
                part_count   = machine_part_count.get(m, max_part_count)
                count_score  = part_count / max_part_count
                co_penalty   = (co_hrs / AVAILABLE_HOURS) * 0.3
                util_penalty = (used / AVAILABLE_HOURS) * 0.2
                same_part_bonus = -0.15 if (last == part) else 0.0
                cost = (count_score + co_penalty + util_penalty
                        + same_part_bonus + color_bonus)
                ranked.append((m, co_hrs, effective_free, cost))
            ranked.sort(key=lambda x: (round(x[3], 8), x[0]))
            return ranked

        def rank_machines(part, machines_to_try, machine_hours,
                          machine_last_part, inv_days,
                          exclude_fixed_machines=True,
                          allow_fixed_overflow=False):
            category  = part_category.get(part, "Stranger")
            new_color = part_color.get(part, "UNKNOWN")
            fixed_m   = part_fixed_machine.get(part)
            is_fixed  = fixed_m is not None
            if is_fixed:
                runner_lock = False
            else:
                runner_lock = (category == "Runner" and inv_days < RUNNER_PRIORITY_DAYS)

            effective_candidates = []
            for m in sorted(machines_to_try):
                # V13: fixed machines are NEVER available to non-fixed parts
                if m in machine_fixed_parts:
                    if not is_fixed:
                        continue
                    elif m != fixed_m:
                        continue
                effective_candidates.append(m)

            if is_fixed and fixed_m in effective_candidates:
                used_f = machine_hours.get(fixed_m, 0)
                free_f = round(AVAILABLE_HOURS - used_f, 4)
                co_f   = _co_hrs_for(part, fixed_m, machine_last_part)
                eff_f  = round(free_f - co_f, 4)
                if eff_f >= MIN_RUN_HOURS:
                    fallback = _rank_normal(
                        part,
                        [m for m in effective_candidates if m != fixed_m],
                        machine_hours, machine_last_part, new_color, runner_lock
                    )
                    return [(fixed_m, co_f, eff_f, -1.0)] + fallback, runner_lock
                remaining = [m for m in effective_candidates if m != fixed_m]
                return _rank_normal(
                    part, remaining, machine_hours, machine_last_part,
                    new_color, runner_lock
                ), runner_lock

            return _rank_normal(
                part, effective_candidates, machine_hours, machine_last_part,
                new_color, runner_lock
            ), runner_lock

        # =========================================================
        # PLANNED QTY TRACKING HELPERS
        # =========================================================

        def _get_part_total_qty(part, plan):
            return sum(float(r.get("Production_Qty", 0)) for r in plan if r["Part"] == part)

        def _is_indent_met(part, plan):
            eff = effective_daily.get(part, 0)
            if eff <= 0:
                return True
            total_qty = _get_part_total_qty(part, plan)
            return total_qty >= (eff - 0.5)

        def _parts_with_indent_met(plan):
            part_qty = defaultdict(float)
            for row in plan:
                part_qty[row["Part"]] += float(row.get("Production_Qty", 0))
            result = set()
            for p, qty in part_qty.items():
                eff = effective_daily.get(p, 0)
                if eff > 0 and qty >= (eff - 0.5):
                    result.add(p)
            return result

        def _current_co_count(plan):
            return sum(1 for r in plan if r.get("Changeover") == "Yes")

        # =========================================================
        # V13 — FIXED MACHINE SCHEDULING (ABSOLUTE LOCKDOWN)
        # =========================================================
        #
        # Fixed machines ALWAYS run their fixed parts for the full available
        # shift. No phase A/B. No OPD cap. No inventory-level exception.
        # The machine is fully consumed by its fixed parts every day.
        #
        # If a machine has multiple fixed parts, they are scheduled in order
        # of lowest days-coverage first (most urgent part runs first), and
        # hours are split proportionally by effective_daily among them until
        # the shift is exhausted.
        # =========================================================

        def schedule_fixed_machines(machine_hours, machine_last_part,
                                    current_inventory, plan, already_planned,
                                    priority_scores, scenario_id):
            """
            V13: Every fixed machine runs ONLY its fixed parts, ALWAYS, for
            the full shift. Inventory level of fixed parts is irrelevant.
            No phase A/B, no OPD cap, no skipping.
            """
            fixed_plan_rows = []

            for machine in sorted(machine_fixed_parts.keys()):
                fixed_parts = machine_fixed_parts[machine]
                if not fixed_parts:
                    continue

                # Order fixed parts by urgency: lowest days-coverage first
                # so the most critical part runs first within the machine shift
                def _urgency_key(p):
                    eff = effective_daily.get(p, 0)
                    inv = current_inventory.get(p, 0)
                    return (inv / eff if eff > 0 else 999.0, p)

                ordered_parts = sorted(fixed_parts, key=_urgency_key)

                # Calculate total effective_daily weight for proportional split
                total_eff = sum(effective_daily.get(p, 0) for p in ordered_parts)

                machine_budget = AVAILABLE_HOURS  # full shift, no cap

                for idx, chosen in enumerate(ordered_parts):
                    remaining_budget = round(machine_budget - machine_hours.get(machine, 0), 4)
                    if remaining_budget < MIN_RUN_HOURS:
                        already_planned.add(chosen)
                        continue

                    eff     = effective_daily.get(chosen, 0)
                    i_daily = indent_daily.get(chosen, 0)
                    d_daily = demand_daily.get(chosen, 0)
                    r_val   = rate.get(chosen, 1)
                    monthly = indent_monthly.get(chosen, 0)
                    color   = part_color.get(chosen, "UNKNOWN")
                    score   = priority_scores.get(chosen, 0)

                    # Proportional hour allocation based on effective_daily share
                    # Last part in list gets remaining hours
                    if idx < len(ordered_parts) - 1 and total_eff > 0:
                        part_share = eff / total_eff
                        alloc_hrs  = round(machine_budget * part_share, 4)
                        # Ensure at least MIN_RUN_HOURS
                        alloc_hrs  = max(MIN_RUN_HOURS, min(alloc_hrs, remaining_budget))
                    else:
                        # Last part: give all remaining hours
                        alloc_hrs = remaining_budget

                    alloc_hrs = max(MIN_RUN_HOURS, alloc_hrs)
                    qty       = round(alloc_hrs * r_val, 0)

                    machine_hours[machine] = round(
                        machine_hours.get(machine, 0) + alloc_hrs, 4
                    )
                    current_inventory[chosen] = round(
                        current_inventory.get(chosen, 0) + qty, 0
                    )
                    machine_last_part[machine] = chosen
                    already_planned.add(chosen)

                    indent_met = qty >= (eff - 0.5) if eff > 0 else True

                    row = {
                        "Part":     chosen,
                        "Color":    color,
                        "Category": part_category.get(chosen, "Runner"),
                        "Fixed_Machine": machine,
                        "Fixed_Used":    "YES — Lockdown (always runs)",
                        "Machine":       machine,
                        "Run_Hours":     round(alloc_hrs, 3),
                        "Changeover_Hrs":  0.0,
                        "Total_Hrs_Used":  round(alloc_hrs, 3),
                        "Rate_Per_Hour":   round(r_val, 2),
                        "Production_Qty":  qty,
                        "Monthly_Indent":  round(monthly, 0),
                        "Indent_Daily":    round(i_daily, 2),
                        "Demand_Daily":    round(d_daily, 2),
                        "Effective_Daily": round(eff, 2),
                        "Today_Target":    round(today_target_qty.get(chosen, 0), 0),
                        "Changeover":  "No",
                        "Color_Purge": "No",
                        "Type":  "Fixed-Lockdown (no cap, no exception)",
                        "Role":  "Primary",
                        "Tools_Available": tools_available.get(chosen, 1),
                        "Tools_Used":      1,
                        "Runner_Lock":     "No",
                        "Priority_Score":  score,
                        "Phase":           1,
                        "Indent_Met":  "YES" if indent_met else "NO — partial (hours limited by co-fixed parts)",
                        "Demand_Met":  "YES" if qty >= d_daily else "NO — partial",
                        "Stagger_Adjusted": "No",
                    }
                    plan.append(row)
                    fixed_plan_rows.append(row)

            # All fixed machines are fully consumed — mark them as phase_a so
            # downstream passes (utilization enforcer, strategic filler, etc.)
            # completely ignore them
            phase_a_machines = set(machine_fixed_parts.keys())
            return phase_a_machines, fixed_plan_rows

        # =========================================================
        # INV BUILD HELPER
        # =========================================================

        def _do_inv_build(part, machine, scenario_id, machine_hours,
                          current_inventory, rows_to_extend, r_val, eff):
            # V13: never extend on a fixed machine
            if machine in machine_fixed_parts:
                return
            inv_after    = current_inventory.get(part, 0)
            cap_qty      = _opd_cap_qty(part, scenario_id, current_inventory)
            headroom_qty = max(0.0, cap_qty - inv_after)
            if headroom_qty <= 0:
                return
            target_row = None
            if isinstance(rows_to_extend, list):
                for row in rows_to_extend:
                    if isinstance(row, dict) and row.get("Part") == part and row.get("Machine") == machine:
                        target_row = row
                        break
            if target_row is None:
                return
            free_m     = round(AVAILABLE_HOURS - machine_hours.get(machine, 0), 4)
            if free_m < 0.05:
                return
            extend_hrs = min(free_m, headroom_qty / r_val if r_val > 0 else 0)
            if extend_hrs < 0.05:
                return
            extra_qty  = round(extend_hrs * r_val, 0)
            target_row["Run_Hours"]      = round(float(target_row["Run_Hours"]) + extend_hrs, 3)
            target_row["Total_Hrs_Used"] = round(
                float(target_row["Changeover_Hrs"]) + float(target_row["Run_Hours"]), 3)
            target_row["Production_Qty"] = round(float(target_row["Production_Qty"]) + extra_qty, 0)
            target_row["Type"] = str(target_row["Type"]) + "+InvBuild"
            machine_hours[machine]  = round(machine_hours.get(machine, 0) + extend_hrs, 4)
            current_inventory[part] = round(current_inventory.get(part, 0) + extra_qty, 0)

        # =========================================================
        # TOOL-AWARE ASSIGNMENT  (V13 — effective_daily as hard floor)
        # =========================================================

        def assign_part(part, scenario_id, machine_hours, machine_last_part,
                        current_inventory, plan, already_planned, priority_scores):
            eff      = effective_daily.get(part, 0)
            i_daily  = indent_daily.get(part, 0)
            d_daily  = demand_daily.get(part, 0)
            monthly  = indent_monthly.get(part, 0)
            r_val    = rate.get(part, 1)
            inv_now  = current_inventory.get(part, 0)
            category = part_category.get(part, "Stranger")
            tools    = tools_available.get(part, 1)
            score    = priority_scores.get(part, 0)
            color    = part_color.get(part, "UNKNOWN")
            inv_days = inv_now / eff if eff > 0 else 999
            compatible = vt_compat.get(part, [])
            fixed_m    = part_fixed_machine.get(part)

            if not compatible:
                return []

            # V13: production floor = effective_daily (max of indent and demand)
            # shortfall is always measured against this floor
            total_shortfall = max(0.0, eff - inv_now)
            hrs_for_full    = max(MIN_RUN_HOURS, total_shortfall / r_val if r_val > 0 else MIN_RUN_HOURS)

            new_rows        = []
            produced_so_far = 0.0
            tools_used      = 0
            used_machines   = set()

            already_on_fixed = (
                fixed_m is not None
                and any(r["Machine"] == fixed_m and r["Part"] == part for r in plan)
            )

            if already_on_fixed:
                qty_from_fixed = sum(
                    float(r["Production_Qty"])
                    for r in plan if r["Part"] == part and r["Machine"] == fixed_m
                )
                produced_so_far = qty_from_fixed
                tools_used = 1
                used_machines.add(fixed_m)
                indent_met_on_fixed = (produced_so_far >= total_shortfall - 0.5)
                if indent_met_on_fixed:
                    already_planned.add(part)
                    return []
                else:
                    if tools <= 1:
                        already_planned.add(part)
                        return []
            else:
                ranked, runner_lock = rank_machines(
                    part, compatible, machine_hours, machine_last_part, inv_days
                )
                if not ranked:
                    return []
                m1, co1, eff1, _ = ranked[0]
                run1   = min(eff1, hrs_for_full)
                run1   = max(run1, MIN_RUN_HOURS)
                qty1   = round(run1 * r_val, 0)
                fixed_used = (fixed_m is not None and m1 == fixed_m)
                machine_hours[m1]       = round(machine_hours.get(m1, 0) + co1 + run1, 4)
                current_inventory[part] = round(current_inventory.get(part, 0) + qty1, 0)
                machine_last_part[m1]   = part
                produced_so_far        += qty1
                tools_used             += 1
                used_machines.add(m1)
                already_planned.add(part)
                indent_met_on_primary = (produced_so_far >= total_shortfall - 0.5)
                last_m1 = machine_state.get(m1)
                purge_applied = (
                    last_m1 is not None and last_m1 != part
                    and ALL_KNOWN_COLORS.get(last_m1, "UNKNOWN") != color
                    and ALL_KNOWN_COLORS.get(last_m1, "UNKNOWN") not in ("UNKNOWN", "NEEDS_PURGE")
                    and color != "UNKNOWN"
                )
                type_tag = "Primary"
                if inv_now == 0:
                    type_tag += " [ZERO-INV]"
                if fixed_used:
                    type_tag += " [FIXED-MACHINE]"
                elif fixed_m is not None:
                    type_tag += " [FIXED-FALLBACK]"

                new_rows.append({
                    "Part":    part,
                    "Color":   color,
                    "Category": category,
                    "Fixed_Machine": fixed_m or "—",
                    "Fixed_Used":   "YES" if fixed_used else ("FALLBACK" if fixed_m else "N/A"),
                    "Machine":      m1,
                    "Run_Hours":    round(run1, 3),
                    "Changeover_Hrs": round(co1, 3),
                    "Total_Hrs_Used": round(co1 + run1, 3),
                    "Rate_Per_Hour":  round(r_val, 2),
                    "Production_Qty": qty1,
                    "Monthly_Indent": round(monthly, 0),
                    "Indent_Daily":   round(i_daily, 2),
                    "Demand_Daily":   round(d_daily, 2),
                    "Effective_Daily": round(eff, 2),
                    "Today_Target":   round(today_target_qty.get(part, 0), 0),
                    "Changeover":  "No" if co1 == 0 else "Yes",
                    "Color_Purge": "Yes" if purge_applied else "No",
                    "Type": type_tag,
                    "Role": "Primary",
                    "Tools_Available": tools,
                    "Tools_Used": 1,
                    "Runner_Lock": "YES" if runner_lock else "No",
                    "Priority_Score": score,
                    "Phase": 1,
                    "Indent_Met": "YES" if indent_met_on_primary else "NO — shortfall remains",
                    "Demand_Met": "YES" if produced_so_far >= d_daily else "NO — demand shortfall",
                    "Stagger_Adjusted": "No",
                })
                if indent_met_on_primary:
                    _do_inv_build(part, m1, scenario_id, machine_hours,
                                  current_inventory, new_rows, r_val, eff)
                    for row in new_rows:
                        row["Tools_Used"] = tools_used
                    return new_rows

            is_critical = (inv_now == 0)
            if category in ("Stranger", "Repeater"):
                tool_hard_cap = 1
            else:
                tool_hard_cap = tools if is_critical else min(2, tools)

            if tool_hard_cap <= 1 and tools_used >= 1:
                if new_rows:
                    _do_inv_build(part, new_rows[0]["Machine"], scenario_id, machine_hours,
                                  current_inventory, new_rows, r_val, eff)
                for row in new_rows:
                    row["Tools_Used"] = tools_used
                return new_rows

            while (produced_so_far < (total_shortfall - 0.5) and tools_used < tool_hard_cap):
                shortfall_now = total_shortfall - produced_so_far
                hrs_needed    = max(MIN_RUN_HOURS, shortfall_now / r_val if r_val > 0 else MIN_RUN_HOURS)
                remaining_machines = sorted([
                    m for m in compatible
                    if m not in used_machines and m not in machine_fixed_parts
                ])
                ranked_next, _ = rank_machines(
                    part, remaining_machines, machine_hours, machine_last_part, inv_days
                )
                if not ranked_next:
                    break
                mx, cox, effx, _ = ranked_next[0]
                run_x  = max(MIN_RUN_HOURS, min(effx, hrs_needed))
                qty_x  = round(run_x * r_val, 0)
                machine_hours[mx]       = round(machine_hours.get(mx, 0) + cox + run_x, 4)
                current_inventory[part] = round(current_inventory.get(part, 0) + qty_x, 0)
                machine_last_part[mx]   = part
                produced_so_far        += qty_x
                tools_used             += 1
                used_machines.add(mx)
                indent_met_here = (produced_so_far >= total_shortfall - 0.5)
                last_mx = machine_state.get(mx)
                purge_x = (
                    last_mx is not None and last_mx != part
                    and ALL_KNOWN_COLORS.get(last_mx, "UNKNOWN") != color
                    and ALL_KNOWN_COLORS.get(last_mx, "UNKNOWN") not in ("UNKNOWN", "NEEDS_PURGE")
                    and color != "UNKNOWN"
                )
                new_rows.append({
                    "Part":    part,
                    "Color":   color,
                    "Category": category,
                    "Fixed_Machine": fixed_m or "—",
                    "Fixed_Used":    "N/A — expansion",
                    "Machine":       mx,
                    "Run_Hours":     round(run_x, 3),
                    "Changeover_Hrs":  round(cox, 3),
                    "Total_Hrs_Used":  round(cox + run_x, 3),
                    "Rate_Per_Hour":   round(r_val, 2),
                    "Production_Qty":  qty_x,
                    "Monthly_Indent":  round(monthly, 0),
                    "Indent_Daily":    round(i_daily, 2),
                    "Demand_Daily":    round(d_daily, 2),
                    "Effective_Daily": round(eff, 2),
                    "Today_Target":    round(today_target_qty.get(part, 0), 0),
                    "Changeover":  "No" if cox == 0 else "Yes",
                    "Color_Purge": "Yes" if purge_x else "No",
                    "Type": "Tool-Expansion",
                    "Role": f"Tool-Expansion (tool {tools_used}/{tools})",
                    "Tools_Available": tools,
                    "Tools_Used": tools_used,
                    "Runner_Lock": "No",
                    "Priority_Score": score,
                    "Phase": 2,
                    "Indent_Met": "YES" if indent_met_here else "NO — shortfall remains",
                    "Demand_Met": "YES" if produced_so_far >= d_daily else "NO — demand shortfall",
                    "Stagger_Adjusted": "No",
                })

            if new_rows:
                _do_inv_build(part, new_rows[0]["Machine"], scenario_id, machine_hours,
                              current_inventory, new_rows, r_val, eff)
            for row in new_rows:
                row["Tools_Used"] = tools_used
            return new_rows

        # =========================================================
        # RUNNER PRIORITY ENFORCEMENT
        # =========================================================

        def enforce_runner_priority(plan, machine_hours, machine_last_part,
                                    current_inventory, already_planned,
                                    priority_scores, inventory_start_of_day):
            runner_priority_log = []
            critical_runners    = []
            for part in sorted(vt_compat.keys()):
                if part in already_planned:
                    continue
                if part_category.get(part, "Stranger") != "Runner":
                    continue
                eff   = effective_daily.get(part, 0)
                r_val = rate.get(part, 1)
                if eff <= 0 or r_val <= 0:
                    continue
                inv_now  = current_inventory.get(part, 0)
                days_now = inv_now / eff if eff > 0 else 999
                if days_now >= RUNNER_PRIORITY_DAYS:
                    continue
                skip, _ = should_skip(part)
                if skip:
                    continue
                if not vt_compat.get(part):
                    continue
                critical_runners.append(part)

            if not critical_runners:
                return runner_priority_log

            # V13: sort by (effective_daily DESC, inventory_ratio ASC) — highest demand, lowest coverage first
            critical_runners.sort(key=lambda p: (
                -effective_daily.get(p, 0),
                (inventory.get(p, 0) / effective_daily.get(p, 1)) if effective_daily.get(p, 0) > 0 else 999,
                p,
            ))

            for runner in critical_runners:
                r_eff      = effective_daily.get(runner, 0)
                r_i_daily  = indent_daily.get(runner, 0)
                r_d_daily  = demand_daily.get(runner, 0)
                r_inv      = current_inventory.get(runner, 0)
                r_inv_sod  = inventory_start_of_day.get(runner, r_inv)
                r_rate     = rate.get(runner, 1)
                r_color    = part_color.get(runner, "UNKNOWN")
                r_score    = priority_scores.get(runner, 0)
                r_monthly  = indent_monthly.get(runner, 0)
                fixed_m    = part_fixed_machine.get(runner)
                r_days_now = r_inv / r_eff if r_eff > 0 else 0
                r_days_sod = r_inv_sod / r_eff if r_eff > 0 else 0
                shortfall_qty = max(0.0, r_eff - r_inv)
                hours_needed  = max(MIN_RUN_HOURS, shortfall_qty / r_rate if r_rate > 0 else MIN_RUN_HOURS)

                machines_runner_on = {row["Machine"] for row in plan if row["Part"] == runner}
                if len(machines_runner_on) >= tools_available.get(runner, 1):
                    runner_priority_log.append({
                        "Runner_Part": runner,
                        "Runner_Days_SOD": round(r_days_sod, 2),
                        "Runner_Days_Now": round(r_days_now, 2),
                        "Fixed_Machine":   fixed_m or "—",
                        "Fixed_Used":      "FAILED — tool cap",
                        "Runner_Indent_Daily":    round(r_i_daily, 2),
                        "Runner_Demand_Daily":    round(r_d_daily, 2),
                        "Runner_Effective_Daily": round(r_eff, 2),
                        "Runner_Inv_Before_SOD":  round(r_inv_sod, 0),
                        "Runner_Shortfall": round(shortfall_qty, 0),
                        "Machine_Assigned": "—",
                        "Hours_Needed": round(hours_needed, 3),
                        "Hours_Assigned": 0, "Qty_Produced": 0,
                        "Displacement_Used": "N/A", "Victims": "—",
                        "Total_Hours_Reclaimed": "—",
                        "Result": "FAILED — all tools deployed",
                    })
                    continue

                # V13: runner priority only uses non-fixed machines
                compatible_machines = sorted([
                    m for m in vt_compat.get(runner, [])
                    if m not in machine_fixed_parts
                ])

                best_machine      = None
                best_co_hrs       = 0.0
                best_victims      = []
                best_disruption   = float("inf")
                best_machine_name = ""

                for m in compatible_machines:
                    co_hrs   = _co_hrs_for(runner, m, machine_last_part)
                    used_hrs = machine_hours.get(m, 0)
                    free_hrs = round(AVAILABLE_HOURS - used_hrs, 4)
                    eff_free = round(free_hrs - co_hrs, 4)

                    if eff_free >= hours_needed:
                        run_h = max(MIN_RUN_HOURS, min(eff_free, hours_needed))
                        qty   = round(run_h * r_rate, 0)
                        machine_hours[m]          = round(used_hrs + co_hrs + run_h, 4)
                        current_inventory[runner] = round(current_inventory.get(runner, 0) + qty, 0)
                        machine_last_part[m]      = runner
                        already_planned.add(runner)
                        purge = co_hrs > vt_changeover.get(m, DEFAULT_CHANGEOVER_HRS) + 0.001
                        plan.append({
                            "Part":    runner, "Color": r_color, "Category": "Runner",
                            "Fixed_Machine": fixed_m or "—",
                            "Fixed_Used": "N/A",
                            "Machine":       m,
                            "Run_Hours":     round(run_h, 3),
                            "Changeover_Hrs": round(co_hrs, 3),
                            "Total_Hrs_Used": round(co_hrs + run_h, 3),
                            "Rate_Per_Hour":  round(r_rate, 2),
                            "Production_Qty": qty,
                            "Monthly_Indent":    round(r_monthly, 0),
                            "Indent_Daily":      round(r_i_daily, 2),
                            "Demand_Daily":      round(r_d_daily, 2),
                            "Effective_Daily":   round(r_eff, 2),
                            "Today_Target":      round(today_target_qty.get(runner, 0), 0),
                            "Changeover":  "No" if co_hrs == 0 else "Yes",
                            "Color_Purge": "Yes" if purge else "No",
                            "Type": "Runner-Priority (free capacity)",
                            "Role": "Primary",
                            "Tools_Available": tools_available.get(runner, 1),
                            "Tools_Used": 1, "Runner_Lock": "No",
                            "Priority_Score": r_score, "Phase": 1,
                            "Indent_Met": "YES" if qty >= shortfall_qty else "NO",
                            "Demand_Met": "YES" if qty >= r_d_daily else "NO",
                            "Stagger_Adjusted": "No",
                        })
                        runner_priority_log.append({
                            "Runner_Part": runner,
                            "Runner_Days_SOD": round(r_days_sod, 2),
                            "Runner_Days_Now": round(r_days_now, 2),
                            "Fixed_Machine": fixed_m or "—",
                            "Fixed_Used": "N/A",
                            "Runner_Indent_Daily":    round(r_i_daily, 2),
                            "Runner_Demand_Daily":    round(r_d_daily, 2),
                            "Runner_Effective_Daily": round(r_eff, 2),
                            "Runner_Inv_Before_SOD":  round(r_inv_sod, 0),
                            "Runner_Shortfall": round(shortfall_qty, 0),
                            "Machine_Assigned": m,
                            "Hours_Needed": round(hours_needed, 3),
                            "Hours_Assigned": round(run_h, 3),
                            "Qty_Produced": qty,
                            "Displacement_Used": "No — free capacity", "Victims": "—",
                            "Total_Hours_Reclaimed": "—",
                            "Result": "PLANNED — free capacity",
                        })
                        best_machine = "DONE"
                        break

                    machine_rows = [r for r in plan if r["Machine"] == m]
                    yieldable    = [
                        r for r in machine_rows
                        if part_category.get(r["Part"], "Stranger") in ("Stranger", "Repeater")
                        and (current_inventory.get(r["Part"], 0) / effective_daily.get(r["Part"], 1)
                             if effective_daily.get(r["Part"], 0) > 0 else 999) > r_days_now
                    ]
                    if not yieldable:
                        continue
                    yieldable.sort(key=lambda r: (
                        effective_daily.get(r["Part"], 0),
                        -float(r.get("Production_Qty", 0)),
                        r["Part"]
                    ))
                    reclaimable_detail = []
                    for row in yieldable:
                        row_run = float(row.get("Run_Hours", 0))
                        if row_run <= 0:
                            continue
                        if row_run > MIN_RUN_HOURS:
                            reclaimable_detail.append((row, round(row_run - MIN_RUN_HOURS, 4), "partial"))
                        else:
                            reclaimable_detail.append((row, round(row_run, 4), "full_remove"))
                    total_reclaimable        = sum(x[1] for x in reclaimable_detail)
                    runner_eff_after_reclaim = round(free_hrs + total_reclaimable - co_hrs, 4)
                    if runner_eff_after_reclaim < max(MIN_RUN_HOURS, hours_needed):
                        continue
                    disruption = total_reclaimable
                    if disruption < best_disruption or (disruption == best_disruption and m < best_machine_name):
                        best_disruption  = disruption
                        best_machine     = m
                        best_co_hrs      = co_hrs
                        best_victims     = reclaimable_detail
                        best_machine_name = m

                if best_machine is None:
                    runner_priority_log.append({
                        "Runner_Part": runner,
                        "Runner_Days_SOD": round(r_days_sod, 2),
                        "Runner_Days_Now": round(r_days_now, 2),
                        "Fixed_Machine": fixed_m or "—",
                        "Fixed_Used": "FAILED",
                        "Runner_Indent_Daily":    round(r_i_daily, 2),
                        "Runner_Demand_Daily":    round(r_d_daily, 2),
                        "Runner_Effective_Daily": round(r_eff, 2),
                        "Runner_Inv_Before_SOD":  round(r_inv_sod, 0),
                        "Runner_Shortfall": round(shortfall_qty, 0),
                        "Machine_Assigned": "—",
                        "Hours_Needed": round(hours_needed, 3),
                        "Hours_Assigned": 0, "Qty_Produced": 0,
                        "Displacement_Used": "N/A", "Victims": "—",
                        "Total_Hours_Reclaimed": "—",
                        "Result": "FAILED — no eligible machine",
                    })
                    continue
                if best_machine == "DONE":
                    continue

                hours_to_free    = hours_needed
                victim_log_parts = []
                total_reclaimed  = 0.0
                for (victim_row, reclaimable_hrs, reclaim_type) in best_victims:
                    if hours_to_free <= 0.001:
                        break
                    vpart      = victim_row["Part"]
                    v_rate     = rate.get(vpart, 1)
                    v_run_orig = float(victim_row.get("Run_Hours", 0))
                    carve_hrs  = round(min(reclaimable_hrs, hours_to_free), 4)
                    if carve_hrs <= 0:
                        continue
                    new_run_hrs = round(v_run_orig - carve_hrs, 4)
                    if new_run_hrs < MIN_RUN_HOURS:
                        plan.remove(victim_row)
                        machine_hours[best_machine] = round(machine_hours.get(best_machine, 0) - v_run_orig, 4)
                        lost_qty = round(v_run_orig * v_rate, 0)
                        current_inventory[vpart] = round(current_inventory.get(vpart, 0) - lost_qty, 0)
                        actually_freed = v_run_orig
                        victim_log_parts.append(f"{vpart} REMOVED")
                    else:
                        lost_qty = round(carve_hrs * v_rate, 0)
                        new_qty  = round(new_run_hrs * v_rate, 0)
                        victim_row["Run_Hours"]      = new_run_hrs
                        victim_row["Production_Qty"] = new_qty
                        victim_row["Total_Hrs_Used"] = round(
                            float(victim_row.get("Changeover_Hrs", 0)) + new_run_hrs, 3)
                        victim_row["Type"] = str(victim_row.get("Type", "")) + " [YIELDED]"
                        machine_hours[best_machine] = round(machine_hours.get(best_machine, 0) - carve_hrs, 4)
                        current_inventory[vpart] = round(current_inventory.get(vpart, 0) - lost_qty, 0)
                        actually_freed = carve_hrs
                        victim_log_parts.append(f"{vpart} -{carve_hrs:.2f}h")
                    hours_to_free   = round(hours_to_free - actually_freed, 4)
                    total_reclaimed = round(total_reclaimed + actually_freed, 4)

                used_now = machine_hours.get(best_machine, 0)
                free_now = round(AVAILABLE_HOURS - used_now, 4)
                eff_free = round(free_now - best_co_hrs, 4)
                if eff_free < MIN_RUN_HOURS:
                    runner_priority_log.append({
                        "Runner_Part": runner,
                        "Runner_Days_SOD": round(r_days_sod, 2),
                        "Runner_Days_Now": round(r_days_now, 2),
                        "Fixed_Machine": fixed_m or "—",
                        "Fixed_Used": "FAILED",
                        "Runner_Indent_Daily":    round(r_i_daily, 2),
                        "Runner_Demand_Daily":    round(r_d_daily, 2),
                        "Runner_Effective_Daily": round(r_eff, 2),
                        "Runner_Inv_Before_SOD":  round(r_inv_sod, 0),
                        "Runner_Shortfall": round(shortfall_qty, 0),
                        "Machine_Assigned": best_machine,
                        "Hours_Needed": round(hours_needed, 3),
                        "Hours_Assigned": 0, "Qty_Produced": 0,
                        "Displacement_Used": "Yes",
                        "Victims": "; ".join(victim_log_parts),
                        "Total_Hours_Reclaimed": round(total_reclaimed, 3),
                        "Result": "FAILED — safety check after carve",
                    })
                    continue

                run_hrs = max(MIN_RUN_HOURS, min(eff_free, hours_needed))
                qty     = round(run_hrs * r_rate, 0)
                purge   = best_co_hrs > vt_changeover.get(best_machine, DEFAULT_CHANGEOVER_HRS) + 0.001
                machine_hours[best_machine]     = round(used_now + best_co_hrs + run_hrs, 4)
                current_inventory[runner]       = round(current_inventory.get(runner, 0) + qty, 0)
                machine_last_part[best_machine] = runner
                already_planned.add(runner)

                plan.append({
                    "Part":    runner, "Color": r_color, "Category": "Runner",
                    "Fixed_Machine": fixed_m or "—",
                    "Fixed_Used": "N/A",
                    "Machine":       best_machine,
                    "Run_Hours":     round(run_hrs, 3),
                    "Changeover_Hrs":  round(best_co_hrs, 3),
                    "Total_Hrs_Used":  round(best_co_hrs + run_hrs, 3),
                    "Rate_Per_Hour":   round(r_rate, 2),
                    "Production_Qty":  qty,
                    "Monthly_Indent":     round(r_monthly, 0),
                    "Indent_Daily":       round(r_i_daily, 2),
                    "Demand_Daily":       round(r_d_daily, 2),
                    "Effective_Daily":    round(r_eff, 2),
                    "Today_Target":       round(today_target_qty.get(runner, 0), 0),
                    "Changeover":  "No" if best_co_hrs == 0 else "Yes",
                    "Color_Purge": "Yes" if purge else "No",
                    "Type": "Runner-Priority [DISPLACED]",
                    "Role": "Primary",
                    "Tools_Available": tools_available.get(runner, 1),
                    "Tools_Used": 1, "Runner_Lock": "No",
                    "Priority_Score": r_score, "Phase": 1,
                    "Indent_Met": "YES" if qty >= shortfall_qty else "NO",
                    "Demand_Met": "YES" if qty >= r_d_daily else "NO",
                    "Stagger_Adjusted": "No",
                })
                runner_priority_log.append({
                    "Runner_Part": runner,
                    "Runner_Days_SOD": round(r_days_sod, 2),
                    "Runner_Days_Now": round(r_days_now, 2),
                    "Fixed_Machine": fixed_m or "—",
                    "Fixed_Used": "N/A",
                    "Runner_Indent_Daily":    round(r_i_daily, 2),
                    "Runner_Demand_Daily":    round(r_d_daily, 2),
                    "Runner_Effective_Daily": round(r_eff, 2),
                    "Runner_Inv_Before_SOD":  round(r_inv_sod, 0),
                    "Runner_Shortfall": round(shortfall_qty, 0),
                    "Machine_Assigned": best_machine,
                    "Hours_Needed": round(hours_needed, 3),
                    "Hours_Assigned": round(run_hrs, 3),
                    "Qty_Produced": qty,
                    "Displacement_Used": "Yes",
                    "Victims": "; ".join(victim_log_parts),
                    "Total_Hours_Reclaimed": round(total_reclaimed, 3),
                    "Result": (
                        "PLANNED — displacement successful"
                        if qty >= shortfall_qty - 0.5 else "PLANNED — partial"
                    ),
                })

            return runner_priority_log

        # =========================================================
        # DISPLACEMENT PRE-PASS
        # =========================================================

        def displace_for_zero_inv(part, machine_hours, machine_last_part,
                                   current_inventory, plan, already_planned, priority_scores,
                                   phase_a_machines):
            eff        = effective_daily.get(part, 0)
            r_val      = rate.get(part, 1)
            category   = part_category.get(part, "Stranger")
            score      = priority_scores.get(part, 0)
            compatible = vt_compat.get(part, [])
            fixed_m    = part_fixed_machine.get(part)
            if not compatible:
                return False
            # V13: never displace on a fixed machine
            candidate_machines = sorted([
                m for m in compatible
                if m not in phase_a_machines
                and m not in machine_fixed_parts
                and round(AVAILABLE_HOURS - machine_hours.get(m, 0), 4) < MIN_RUN_HOURS
            ])
            if not candidate_machines:
                return False

            best_machine     = None
            best_victim_row  = None
            best_victim_days = -1
            best_victim_part = ""
            part_days_cov    = current_inventory.get(part, 0) / eff if eff > 0 else 0

            for m in candidate_machines:
                for row in sorted([r for r in plan if r["Machine"] == m], key=lambda r: r["Part"]):
                    vpart  = row["Part"]
                    v_eff  = effective_daily.get(vpart, 0)
                    vinv   = current_inventory.get(vpart, 0)
                    vdays  = vinv / v_eff if v_eff > 0 else 999
                    if vdays <= part_days_cov:
                        continue
                    if vdays < SAFETY_DAYS or vinv <= 0:
                        continue
                    vrun   = float(row.get("Run_Hours", 0))
                    if vrun - MIN_RUN_HOURS < MIN_RUN_HOURS:
                        continue
                    if vdays > best_victim_days or (vdays == best_victim_days and vpart < best_victim_part):
                        best_victim_days = vdays
                        best_victim_row  = row
                        best_machine     = m
                        best_victim_part = vpart

            if best_machine is None or best_victim_row is None:
                return False

            vpart      = best_victim_row["Part"]
            vr_val     = rate.get(vpart, 1)
            reduce_h   = MIN_RUN_HOURS
            lost_qty   = round(reduce_h * vr_val, 0)
            best_victim_row["Run_Hours"]      = round(float(best_victim_row["Run_Hours"]) - reduce_h, 3)
            best_victim_row["Production_Qty"] = round(float(best_victim_row["Production_Qty"]) - lost_qty, 0)
            best_victim_row["Total_Hrs_Used"] = round(
                float(best_victim_row.get("Changeover_Hrs", 0)) + float(best_victim_row["Run_Hours"]), 3)
            best_victim_row["Type"] = str(best_victim_row.get("Type", "")) + " [DISPLACED]"
            current_inventory[vpart]    = round(current_inventory.get(vpart, 0) - lost_qty, 0)
            machine_hours[best_machine] = round(machine_hours.get(best_machine, 0) - reduce_h, 4)

            p_col = part_color.get(part, "UNKNOWN")
            last  = machine_last_part.get(best_machine)
            l_col = ALL_KNOWN_COLORS.get(last, "UNKNOWN") if last else "UNKNOWN"
            if last is None or last == part:
                co_hrs = 0.0
            else:
                base_co = vt_changeover.get(best_machine, DEFAULT_CHANGEOVER_HRS)
                purge   = COLOR_PURGE_HRS if p_col != l_col and p_col != "UNKNOWN" and l_col not in ("UNKNOWN", "NEEDS_PURGE") else 0.0
                co_hrs  = base_co + purge

            eff_free       = round(AVAILABLE_HOURS - machine_hours.get(best_machine, 0) - co_hrs, 4)
            hrs_for_indent = max(MIN_RUN_HOURS, eff / r_val if r_val > 0 else MIN_RUN_HOURS)
            run_hrs        = max(MIN_RUN_HOURS, min(eff_free, hrs_for_indent))
            qty            = round(run_hrs * r_val, 0)

            machine_hours[best_machine]     = round(machine_hours.get(best_machine, 0) + co_hrs + run_hrs, 4)
            current_inventory[part]         = round(current_inventory.get(part, 0) + qty, 0)
            machine_last_part[best_machine] = part
            already_planned.add(part)

            has_purge  = co_hrs > vt_changeover.get(best_machine, DEFAULT_CHANGEOVER_HRS) + 0.001
            fixed_used = (fixed_m is not None and best_machine == fixed_m)
            i_daily    = indent_daily.get(part, 0)
            d_daily    = demand_daily.get(part, 0)

            plan.append({
                "Part":    part, "Color": p_col, "Category": category,
                "Fixed_Machine": fixed_m or "—",
                "Fixed_Used": "YES" if fixed_used else ("FALLBACK" if fixed_m else "N/A"),
                "Machine":       best_machine,
                "Run_Hours":     round(run_hrs, 3),
                "Changeover_Hrs": round(co_hrs, 3),
                "Total_Hrs_Used": round(co_hrs + run_hrs, 3),
                "Rate_Per_Hour":  round(r_val, 2),
                "Production_Qty": qty,
                "Monthly_Indent": round(indent_monthly.get(part, 0), 0),
                "Indent_Daily":   round(i_daily, 2),
                "Demand_Daily":   round(d_daily, 2),
                "Effective_Daily": round(eff, 2),
                "Today_Target":   round(today_target_qty.get(part, 0), 0),
                "Changeover":  "No" if co_hrs == 0 else "Yes",
                "Color_Purge": "Yes" if has_purge else "No",
                "Type": "Displacement [ZERO-INV]",
                "Role": "Primary",
                "Tools_Available": tools_available.get(part, 1),
                "Tools_Used": 1, "Runner_Lock": "No",
                "Priority_Score": score, "Phase": 1,
                "Indent_Met": "YES" if qty >= eff else "NO",
                "Demand_Met": "YES" if qty >= d_daily else "NO",
                "Stagger_Adjusted": "No",
            })
            return True

        # =========================================================
        # TOOL-CHANGER / STAGGER HELPERS
        # =========================================================

        def _collect_co_events(plan, machines):
            events = []
            for m in sorted(machines):
                m_rows = [r for r in plan if r["Machine"] == m]
                if len(m_rows) < 2:
                    continue
                cursor = 0.0
                for i, row in enumerate(m_rows):
                    co_h  = float(row.get("Changeover_Hrs") or 0.0)
                    run_h = float(row.get("Run_Hours") or 0.0)
                    if co_h > 0 and i > 0:
                        events.append({
                            "machine": m, "part_before": m_rows[i-1]["Part"],
                            "part_after": row["Part"], "co_duration": co_h,
                            "natural_start": cursor, "row_before": m_rows[i-1],
                            "row_after": row, "actual_start": None, "wait_hrs": 0.0,
                        })
                    cursor += co_h + run_h
            return events

        def _recompute_natural_start(ev, plan):
            m, target = ev["machine"], ev["row_after"]
            cursor = 0.0
            for r in [r for r in plan if r["Machine"] == m]:
                if r is target:
                    break
                cursor += float(r.get("Changeover_Hrs") or 0) + float(r.get("Run_Hours") or 0)
            return cursor

        def _machine_spare(m, plan):
            used = sum(float(r.get("Changeover_Hrs") or 0) + float(r.get("Run_Hours") or 0)
                       for r in plan if r["Machine"] == m)
            return max(0.0, AVAILABLE_HOURS - used)

        def _extend_row_before(ev, wait_hrs, plan, machine_hours):
            m     = ev["machine"]
            spare = _machine_spare(m, plan)
            extend_by = min(wait_hrs, spare)
            if extend_by <= 0:
                return 0.0, 0
            rb        = ev["row_before"]
            r_val     = rate.get(rb["Part"], 1.0)
            extra_qty = round(extend_by * r_val, 0)
            rb["Run_Hours"]      = round(float(rb.get("Run_Hours") or 0) + extend_by, 3)
            rb["Production_Qty"] = round(float(rb.get("Production_Qty") or 0) + extra_qty, 0)
            rb["Total_Hrs_Used"] = round(
                float(rb.get("Changeover_Hrs") or 0) + float(rb["Run_Hours"]), 3)
            rb["Stagger_Adjusted"] = "Yes"
            machine_hours[m] = round(machine_hours.get(m, 0) + extend_by, 4)
            return extend_by, int(extra_qty)

        def stagger_changeovers(plan, machines, machine_hours):
            events = _collect_co_events(plan, machines)
            if not events:
                return
            events.sort(key=lambda e: (e["natural_start"], e["machine"]))
            if len(events) > MAX_DAILY_CO:
                events = events[:MAX_DAILY_CO]
            tool_changer_free_at = 0.0
            for idx, ev in enumerate(events, 1):
                natural_start = _recompute_natural_start(ev, plan)
                co_h          = ev["co_duration"]
                actual_start  = max(natural_start, tool_changer_free_at)
                wait_hrs      = round(actual_start - natural_start, 4)
                tool_changer_free_at = actual_start + co_h
                if wait_hrs > 0.001:
                    _extend_row_before(ev, wait_hrs, plan, machine_hours)
                ev["actual_start"] = actual_start
                ev["wait_hrs"]     = wait_hrs

        # =========================================================
        # EXTEND EXISTING ON MACHINE HELPER
        # =========================================================

        def _extend_existing_on_machine(m, plan, machine_hours, current_inventory,
                                         priority_scores, ceiling_days, remaining):
            # V13: never extend on a fixed machine
            if m in machine_fixed_parts:
                return 0.0, remaining
            consumed   = 0.0
            parts_on_m = sorted(
                [row for row in plan if row["Machine"] == m],
                key=lambda r: (
                    -(priority_scores.get(r["Part"], 0) if isinstance(priority_scores.get(r["Part"], 0), (int, float)) else 0),
                    r["Part"]
                ),
            )
            for row in parts_on_m:
                if remaining < 0.001:
                    break
                p_ext    = row["Part"]
                if is_hard_skip(p_ext):
                    continue
                r_ext    = rate.get(p_ext, 1)
                eff_p    = effective_daily.get(p_ext, 0)
                inv_now  = current_inventory.get(p_ext, 0)
                headroom = max(0.0, ceiling_days * eff_p - inv_now) if eff_p > 0 else 0
                ext_hrs  = min(remaining, headroom / r_ext if r_ext > 0 else 0)
                if ext_hrs < 0.001:
                    continue
                extra_qty = round(ext_hrs * r_ext, 0)
                row["Run_Hours"]      = round(float(row.get("Run_Hours", 0)) + ext_hrs, 3)
                row["Production_Qty"] = round(float(row.get("Production_Qty", 0)) + extra_qty, 0)
                row["Total_Hrs_Used"] = round(float(row.get("Changeover_Hrs", 0)) + float(row["Run_Hours"]), 3)
                row["Type"] = str(row.get("Type", "")) + f"+Ext{ceiling_days}d"
                machine_hours[m]         = round(machine_hours.get(m, 0) + ext_hrs, 4)
                current_inventory[p_ext] = round(current_inventory.get(p_ext, 0) + extra_qty, 0)
                remaining = round(remaining - ext_hrs, 4)
                consumed += ext_hrs
            return consumed, remaining

        # =========================================================
        # ADD PART TO MACHINE HELPER
        # =========================================================

        def _add_part_to_machine(p, m, run_hrs, co_hrs, machine_hours,
                                  machine_last_part, current_inventory,
                                  already_planned, plan, priority_scores,
                                  type_label, role_label):
            # V13: never add to a fixed machine
            if m in machine_fixed_parts:
                return 0
            r_val   = rate.get(p, 1)
            qty     = round(run_hrs * r_val, 0)
            last    = machine_last_part.get(m)
            p_color = part_color.get(p, "UNKNOWN")
            l_color = ALL_KNOWN_COLORS.get(last, "UNKNOWN") if last else "UNKNOWN"
            has_purge = (last is not None and last != p and p_color != l_color
                         and p_color != "UNKNOWN" and l_color not in ("UNKNOWN", "NEEDS_PURGE"))
            machine_hours[m]    = round(machine_hours.get(m, 0) + co_hrs + run_hrs, 4)
            current_inventory[p] = round(current_inventory.get(p, 0) + qty, 0)
            machine_last_part[m] = p
            already_planned.add(p)
            i_daily = indent_daily.get(p, 0)
            d_daily = demand_daily.get(p, 0)
            eff     = effective_daily.get(p, 0)
            plan.append({
                "Part":    p, "Color": p_color, "Category": part_category.get(p, "Stranger"),
                "Fixed_Machine": part_fixed_machine.get(p, "—"),
                "Fixed_Used":    "N/A — enforcer",
                "Machine":       m,
                "Run_Hours":     round(run_hrs, 3),
                "Changeover_Hrs": round(co_hrs, 3),
                "Total_Hrs_Used": round(co_hrs + run_hrs, 3),
                "Rate_Per_Hour":  round(r_val, 2),
                "Production_Qty": qty,
                "Monthly_Indent": round(indent_monthly.get(p, 0), 0),
                "Indent_Daily":   round(i_daily, 2),
                "Demand_Daily":   round(d_daily, 2),
                "Effective_Daily": round(eff, 2),
                "Today_Target":   round(today_target_qty.get(p, 0), 0),
                "Changeover":  "No" if co_hrs == 0 else "Yes",
                "Color_Purge": "Yes" if has_purge else "No",
                "Type": type_label, "Role": role_label,
                "Tools_Available": tools_available.get(p, 1),
                "Tools_Used": 1, "Runner_Lock": "No",
                "Priority_Score": round(priority_scores.get(p, 0), 2),
                "Phase": 1, "Stagger_Adjusted": "No",
            })
            return qty

        # =========================================================
        # UTILIZATION ENFORCER
        # =========================================================

        def utilization_enforcer(plan, machine_hours, machine_last_part,
                                  all_parts, already_planned,
                                  current_inventory, scenario_id, priority_scores,
                                  phase_a_machines):
            micro_idle_log = []
            floor_hrs      = AVAILABLE_HOURS * (UTIL_TARGET_PCT / 100.0)
            # V13: phase_a_machines includes ALL fixed machines, skip them entirely
            machines_by_util = sorted(
                [m for m in vt_machines if m not in phase_a_machines],
                key=lambda m: (machine_hours.get(m, 0), m)
            )

            for m in machines_by_util:
                remaining  = round(AVAILABLE_HOURS - machine_hours.get(m, 0), 4)
                if remaining < 0.05:
                    continue
                last_on_m  = machine_last_part.get(m)
                last_color = part_color.get(last_on_m, "UNKNOWN") if last_on_m else "UNKNOWN"

                current_util_hrs = machine_hours.get(m, 0)
                if current_util_hrs < floor_hrs:
                    needed = round(floor_hrs - current_util_hrs, 4)
                    _, remaining = _extend_existing_on_machine(
                        m, plan, machine_hours, current_inventory,
                        priority_scores, opd_cap(scenario_id), min(needed, remaining))
                    remaining = round(AVAILABLE_HOURS - machine_hours.get(m, 0), 4)

                if remaining < 0.05:
                    continue

                if remaining >= MIN_RUN_HOURS:
                    machine_compat = machine_compatible_parts.get(m, [])
                    unplanned = []
                    for p in machine_compat:
                        if indent_monthly.get(p, 0) <= 0 or rate.get(p, 0) <= 0:
                            continue
                        if terminal_blocked(p)[0]:
                            continue
                        pfm = part_fixed_machine.get(p)
                        if pfm is not None and pfm != m:
                            continue
                        # V13: skip if this part's fixed machine is a different machine
                        if pfm is not None and pfm == m and m in machine_fixed_parts:
                            continue
                        eff_p    = effective_daily.get(p, 0)
                        inv_now  = current_inventory.get(p, 0)
                        cap_qty  = _opd_cap_qty(p, scenario_id, current_inventory)
                        if eff_p > 0 and inv_now >= cap_qty:
                            continue
                        cat = part_category.get(p, "Stranger")
                        if cat == "Runner" and p in already_planned:
                            if len({row["Machine"] for row in plan if row["Part"] == p}) >= tools_available.get(p, 1):
                                continue
                        unplanned.append(p)

                    def _sort_key(p):
                        total_qty_today = _get_part_total_qty(p, plan)
                        is_unplanned    = 0 if total_qty_today == 0 else 1
                        needs_co        = 0 if (last_on_m is None or last_on_m == p) else 1
                        p_col           = part_color.get(p, "UNKNOWN")
                        same_col        = 0 if (needs_co == 1 and p_col == last_color and p_col != "UNKNOWN") else 1
                        cat_pri         = {"Runner": 0, "Repeater": 1, "Stranger": 2}.get(part_category.get(p, "Stranger"), 2)
                        inv_now_p       = current_inventory.get(p, 0)
                        eff_p           = effective_daily.get(p, 0)
                        days_cov        = inv_now_p / eff_p if eff_p > 0 else 999
                        sc              = priority_scores.get(p, 0)
                        # V13: also sort by effective_daily desc (high demand first)
                        eff_weight      = -eff_p
                        return (is_unplanned, needs_co, same_col, cat_pri, days_cov, eff_weight, -sc, p)

                    unplanned.sort(key=_sort_key)

                    for p in unplanned:
                        if remaining < MIN_RUN_HOURS:
                            break
                        co_hrs   = _co_hrs_for(p, m, machine_last_part)
                        eff_free = round(remaining - co_hrs, 4)
                        if eff_free < MIN_RUN_HOURS:
                            continue
                        if co_hrs > 0 and _current_co_count(plan) >= MAX_DAILY_CO:
                            continue
                        eff_p     = effective_daily.get(p, 0)
                        r_val     = rate.get(p, 1)
                        inv_now_p = current_inventory.get(p, 0)
                        cap_qty   = _opd_cap_qty(p, scenario_id, current_inventory)
                        headroom  = max(0.0, cap_qty - inv_now_p)
                        if headroom <= 0:
                            continue
                        shortfall = max(0.0, eff_p - inv_now_p)
                        min_run   = max(MIN_RUN_HOURS, shortfall / r_val if r_val > 0 else MIN_RUN_HOURS)
                        run_hrs   = max(min_run, min(eff_free, headroom / r_val if r_val > 0 else eff_free))
                        was_unplanned = _get_part_total_qty(p, plan) == 0
                        _add_part_to_machine(
                            p, m, run_hrs, co_hrs, machine_hours, machine_last_part,
                            current_inventory, already_planned, plan, priority_scores,
                            "Filler-Unplanned" if was_unplanned else "Filler-Expand", "Primary",
                        )
                        remaining  = round(remaining - co_hrs - run_hrs, 4)
                        last_on_m  = machine_last_part.get(m)
                        last_color = part_color.get(last_on_m, "UNKNOWN") if last_on_m else "UNKNOWN"

                if remaining < 0.05:
                    continue

                for ceiling in [opd_cap(scenario_id), STRATEGIC_BUFFER_DAYS, ABSOLUTE_MAX_DAYS]:
                    if remaining < 0.001:
                        break
                    _, remaining = _extend_existing_on_machine(
                        m, plan, machine_hours, current_inventory,
                        priority_scores, ceiling, remaining)
                    remaining = round(AVAILABLE_HOURS - machine_hours.get(m, 0), 4)

                if remaining < MIN_RUN_HOURS:
                    if 0.001 < remaining:
                        _, remaining = _extend_existing_on_machine(
                            m, plan, machine_hours, current_inventory,
                            priority_scores, ABSOLUTE_MAX_DAYS, remaining)
                    continue

                indent_met_parts = _parts_with_indent_met(plan)
                runner_spare = sorted([
                    p for p in already_planned
                    if m in vt_compat.get(p, [])
                    and part_category.get(p, "Stranger") == "Runner"
                    and p not in indent_met_parts
                    and m not in [row["Machine"] for row in plan if row["Part"] == p]
                    and len({row["Machine"] for row in plan if row["Part"] == p}) < tools_available.get(p, 1)
                    and (part_fixed_machine.get(p) is None or part_fixed_machine.get(p) == m)
                    and m not in machine_fixed_parts  # V13: not a fixed machine
                ])
                runner_spare.sort(key=lambda p: (
                    0 if (last_on_m is None or last_on_m == p) else 1,
                    0 if (part_color.get(p, "UNKNOWN") == last_color and last_color != "UNKNOWN") else 1,
                    -effective_daily.get(p, 0),   # V13: high demand first
                    -priority_scores.get(p, 0),
                    p,
                ))
                for p in runner_spare:
                    if remaining < MIN_RUN_HOURS:
                        break
                    co_hrs   = _co_hrs_for(p, m, machine_last_part)
                    eff_free = round(remaining - co_hrs, 4)
                    if eff_free < MIN_RUN_HOURS:
                        continue
                    if co_hrs > 0 and _current_co_count(plan) >= MAX_DAILY_CO:
                        continue
                    eff_p     = effective_daily.get(p, 0)
                    r_val     = rate.get(p, 1)
                    inv_now_p = current_inventory.get(p, 0)
                    cap_qty   = _opd_cap_qty(p, scenario_id, current_inventory)
                    headroom  = max(0.0, cap_qty - inv_now_p)
                    if headroom <= 0:
                        continue
                    run_hrs   = max(MIN_RUN_HOURS, min(eff_free, headroom / r_val if r_val > 0 else eff_free))
                    qty       = round(run_hrs * r_val, 0)
                    machine_hours[m]     = round(machine_hours.get(m, 0) + co_hrs + run_hrs, 4)
                    current_inventory[p] = round(current_inventory.get(p, 0) + qty, 0)
                    machine_last_part[m] = p
                    remaining = round(remaining - co_hrs - run_hrs, 4)
                    tools_used_now = len({row["Machine"] for row in plan if row["Part"] == p}) + 1
                    i_daily = indent_daily.get(p, 0)
                    d_daily = demand_daily.get(p, 0)
                    plan.append({
                        "Part":    p, "Color": part_color.get(p, "UNKNOWN"), "Category": "Runner",
                        "Fixed_Machine": part_fixed_machine.get(p, "—"),
                        "Fixed_Used":    "N/A — enforcer",
                        "Machine":       m,
                        "Run_Hours":     round(run_hrs, 3),
                        "Changeover_Hrs": round(co_hrs, 3),
                        "Total_Hrs_Used": round(co_hrs + run_hrs, 3),
                        "Rate_Per_Hour":  round(r_val, 2),
                        "Production_Qty": qty,
                        "Monthly_Indent": round(indent_monthly.get(p, 0), 0),
                        "Indent_Daily":   round(i_daily, 2),
                        "Demand_Daily":   round(d_daily, 2),
                        "Effective_Daily": round(eff_p, 2),
                        "Today_Target":   round(today_target_qty.get(p, 0), 0),
                        "Changeover":  "No" if co_hrs == 0 else "Yes",
                        "Color_Purge": "No",
                        "Type": "Runner-Rerun",
                        "Role": f"Tool-Expansion (tool {tools_used_now})",
                        "Tools_Available": tools_available.get(p, 1),
                        "Tools_Used":     tools_used_now,
                        "Runner_Lock": "No",
                        "Priority_Score": round(priority_scores.get(p, 0), 2),
                        "Phase": 3, "Stagger_Adjusted": "No",
                    })
                    last_on_m  = p
                    last_color = part_color.get(p, "UNKNOWN")
                    break

                if remaining < MIN_RUN_HOURS:
                    remaining = round(AVAILABLE_HOURS - machine_hours.get(m, 0), 4)
                    if 0.001 < remaining:
                        _, remaining = _extend_existing_on_machine(
                            m, plan, machine_hours, current_inventory,
                            priority_scores, ABSOLUTE_MAX_DAYS, remaining)
                    continue

                s4_cands = []
                machine_compat = machine_compatible_parts.get(m, [])
                for p in machine_compat:
                    if terminal_blocked(p)[0]:
                        continue
                    if rate.get(p, 0) <= 0 or indent_monthly.get(p, 0) <= 0:
                        continue
                    pfm = part_fixed_machine.get(p)
                    if pfm is not None and pfm != m:
                        continue
                    if m in machine_fixed_parts:
                        continue
                    s4_cands.append(p)

                s4_cands.sort(key=lambda p: (
                    0 if (last_on_m is None or last_on_m == p) else 1,
                    {"Runner": 0, "Repeater": 1, "Stranger": 2}.get(part_category.get(p, "Stranger"), 2),
                    -effective_daily.get(p, 0),
                    p,
                ))
                for p in s4_cands:
                    if remaining < MIN_RUN_HOURS:
                        break
                    co_hrs   = _co_hrs_for(p, m, machine_last_part)
                    eff_free = round(remaining - co_hrs, 4)
                    if eff_free < MIN_RUN_HOURS:
                        continue
                    if co_hrs > 0 and _current_co_count(plan) >= MAX_DAILY_CO:
                        continue
                    r_val     = rate.get(p, 1)
                    eff_p     = effective_daily.get(p, 0)
                    inv_now_p = current_inventory.get(p, 0)
                    headroom  = (max(0.0, ABSOLUTE_MAX_DAYS * eff_p - inv_now_p)
                                 if eff_p > 0 else (eff_free * r_val))
                    if headroom <= 0:
                        continue
                    run_hrs = max(MIN_RUN_HOURS, min(eff_free, headroom / r_val if r_val > 0 else eff_free))
                    _add_part_to_machine(
                        p, m, run_hrs, co_hrs, machine_hours, machine_last_part,
                        current_inventory, already_planned, plan, priority_scores,
                        "Filler-AnyCompatible", "Primary",
                    )
                    remaining  = round(remaining - co_hrs - run_hrs, 4)
                    last_on_m  = machine_last_part.get(m)
                    last_color = part_color.get(last_on_m, "UNKNOWN") if last_on_m else "UNKNOWN"

                remaining = round(AVAILABLE_HOURS - machine_hours.get(m, 0), 4)
                if 0.001 < remaining:
                    _, remaining = _extend_existing_on_machine(
                        m, plan, machine_hours, current_inventory,
                        priority_scores, ABSOLUTE_MAX_DAYS, remaining)

                final_remaining = round(AVAILABLE_HOURS - machine_hours.get(m, 0), 4)
                if final_remaining >= 0.25:
                    util_final = round((1 - final_remaining / AVAILABLE_HOURS) * 100, 1)
                    micro_idle_log.append({
                        "Machine":         m,
                        "Idle_Hrs":        round(final_remaining, 3),
                        "Utilization_Pct": util_final,
                        "Note": "Exhausted all compatible parts",
                    })

            return micro_idle_log

        # =========================================================
        # STRATEGIC BUFFER FILLER
        # =========================================================

        def strategic_buffer_score(part, current_inventory):
            eff   = effective_daily.get(part, 0.0)
            inv   = current_inventory.get(part, 0.0)
            r_val = rate.get(part, 1.0)
            cat   = part_category.get(part, "Stranger")
            if eff <= 0 or r_val <= 0:
                return 0.0
            velocity = eff / max(float(inv), 1.0)
            cat_w    = {"Runner": 1.0, "Repeater": 0.7, "Stranger": 0.4}.get(cat, 0.4)
            return round(velocity * cat_w * STRATEGIC_PRIORITY_DISCOUNT * 100, 2)

        def strategic_buffer_filler(plan, machine_hours, machine_last_part,
                                    all_parts, already_planned,
                                    current_inventory, scenario_id, phase_a_machines):
            filled_count = 0
            strat_scores = {p: strategic_buffer_score(p, current_inventory) for p in sorted(all_parts)}
            # V13: skip all fixed machines
            machines_by_util = sorted(
                [m for m in vt_machines if m not in phase_a_machines],
                key=lambda m: (machine_hours.get(m, 0), m)
            )

            for m in machines_by_util:
                remaining = round(AVAILABLE_HOURS - machine_hours.get(m, 0), 4)
                if remaining < 0.001:
                    continue
                last_on_m  = machine_last_part.get(m)
                last_color = part_color.get(last_on_m, "UNKNOWN") if last_on_m else "UNKNOWN"

                for ceiling_days in [STRATEGIC_BUFFER_DAYS, ABSOLUTE_MAX_DAYS]:
                    if remaining < 0.001:
                        break
                    consumed, remaining = _extend_existing_on_machine(
                        m, plan, machine_hours, current_inventory,
                        strat_scores, ceiling_days, remaining)
                    if consumed > 0:
                        filled_count += 1
                    remaining = round(AVAILABLE_HOURS - machine_hours.get(m, 0), 4)

                if remaining < MIN_RUN_HOURS:
                    if remaining > 0.001:
                        _, remaining = _extend_existing_on_machine(
                            m, plan, machine_hours, current_inventory,
                            strat_scores, ABSOLUTE_MAX_DAYS, remaining)
                    continue

                machine_compat = machine_compatible_parts.get(m, [])
                candidates     = []
                for p in machine_compat:
                    if terminal_blocked(p)[0]:
                        continue
                    if rate.get(p, 0) <= 0 or indent_monthly.get(p, 0) <= 0:
                        continue
                    pfm = part_fixed_machine.get(p)
                    if pfm is not None and pfm != m:
                        continue
                    if m in machine_fixed_parts:
                        continue
                    cat = part_category.get(p, "Stranger")
                    if cat == "Runner" and p in already_planned:
                        if len({row["Machine"] for row in plan if row["Part"] == p}) >= tools_available.get(p, 1):
                            continue
                    inv_p = current_inventory.get(p, 0)
                    eff_p = effective_daily.get(p, 0)
                    if eff_p > 0 and inv_p >= ABSOLUTE_MAX_DAYS * eff_p:
                        continue
                    candidates.append(p)

                def _strat_sort(p):
                    total_qty_today = _get_part_total_qty(p, plan)
                    is_unplanned    = 0 if total_qty_today == 0 else 1
                    needs_co        = 0 if (last_on_m is None or last_on_m == p) else 1
                    p_col           = part_color.get(p, "UNKNOWN")
                    same_col        = 0 if (needs_co == 1 and p_col == last_color and p_col != "UNKNOWN") else 1
                    sc = strat_scores.get(p, 0)
                    return (is_unplanned, needs_co, same_col, -sc, p)

                candidates.sort(key=_strat_sort)

                for p in candidates:
                    if remaining < MIN_RUN_HOURS:
                        break
                    co_hrs   = _co_hrs_for(p, m, machine_last_part)
                    eff_free = round(remaining - co_hrs, 4)
                    if eff_free < MIN_RUN_HOURS:
                        continue
                    if co_hrs > 0 and _current_co_count(plan) >= MAX_DAILY_CO:
                        continue
                    r_val    = rate.get(p, 1)
                    eff_p    = effective_daily.get(p, 0)
                    inv_now  = current_inventory.get(p, 0)
                    headroom = max(0.0, STRATEGIC_BUFFER_DAYS * eff_p - inv_now)
                    if headroom <= 0:
                        headroom = max(0.0, ABSOLUTE_MAX_DAYS * eff_p - inv_now)
                    if headroom <= 0:
                        continue
                    run_hrs     = max(MIN_RUN_HOURS, min(eff_free, headroom / r_val if r_val > 0 else eff_free))
                    qty         = round(run_hrs * r_val, 0)
                    p_color_val = part_color.get(p, "UNKNOWN")
                    has_purge   = co_hrs > vt_changeover.get(m, DEFAULT_CHANGEOVER_HRS) + 0.001
                    i_daily     = indent_daily.get(p, 0)
                    d_daily     = demand_daily.get(p, 0)

                    machine_hours[m]     = round(machine_hours.get(m, 0) + co_hrs + run_hrs, 4)
                    current_inventory[p] = round(current_inventory.get(p, 0) + qty, 0)
                    machine_last_part[m] = p
                    already_planned.add(p)
                    remaining   = round(remaining - co_hrs - run_hrs, 4)
                    filled_count += 1

                    plan.append({
                        "Part":    p, "Color": p_color_val,
                        "Category": part_category.get(p, "?"),
                        "Fixed_Machine": part_fixed_machine.get(p, "—"),
                        "Fixed_Used":    "N/A — strategic buffer",
                        "Machine":       m,
                        "Run_Hours":     round(run_hrs, 3),
                        "Changeover_Hrs": round(co_hrs, 3),
                        "Total_Hrs_Used": round(co_hrs + run_hrs, 3),
                        "Rate_Per_Hour":  round(r_val, 2),
                        "Production_Qty": qty,
                        "Monthly_Indent": round(indent_monthly.get(p, 0), 0),
                        "Indent_Daily":   round(i_daily, 2),
                        "Demand_Daily":   round(d_daily, 2),
                        "Effective_Daily": round(eff_p, 2),
                        "Today_Target": 0,
                        "Changeover":  "No" if co_hrs == 0 else "Yes",
                        "Color_Purge": "Yes" if has_purge else "No",
                        "Type": "Strategic-Buffer",
                        "Role": "Buffer-Fill",
                        "Tools_Available": tools_available.get(p, 1),
                        "Tools_Used": 1, "Runner_Lock": "No",
                        "Priority_Score": round(strat_scores.get(p, 0), 2),
                        "Phase": 4, "Stagger_Adjusted": "No",
                    })
                    last_on_m  = machine_last_part.get(m)
                    last_color = part_color.get(last_on_m, "UNKNOWN") if last_on_m else "UNKNOWN"

                remaining = round(AVAILABLE_HOURS - machine_hours.get(m, 0), 4)
                if remaining > 0.001:
                    _, remaining = _extend_existing_on_machine(
                        m, plan, machine_hours, current_inventory,
                        strat_scores, ABSOLUTE_MAX_DAYS, remaining)

            return filled_count

        # =========================================================
        # INTRA-MACHINE CO RESEQUENCING
        # =========================================================

        def resequence_machine_rows(plan, machine_last_part_yesterday):
            machine_rows    = defaultdict(list)
            other_rows      = []
            for row in plan:
                m = row.get("Machine")
                if m in vt_machines:
                    machine_rows[m].append(row)
                else:
                    other_rows.append(row)
            resequenced_plan = []
            for m in sorted(vt_machines):
                rows = machine_rows.get(m, [])
                if len(rows) <= 1:
                    resequenced_plan.extend(rows)
                    continue
                # V13: don't resequence fixed machines — their part order is set by urgency key above
                if m in machine_fixed_parts:
                    resequenced_plan.extend(rows)
                    continue
                yesterday_part = machine_last_part_yesterday.get(m)
                ordered   = []
                remaining = list(rows)
                seed = None
                if yesterday_part:
                    for r in remaining:
                        if r["Part"] == yesterday_part:
                            seed = r
                            break
                if seed is None:
                    seed = max(
                        remaining,
                        key=lambda r: (
                            float(r.get("Priority_Score", 0) or 0),
                            -sum(ord(c) for c in str(r.get("Part", "")))
                        )
                    )
                ordered.append(seed)
                remaining.remove(seed)
                while remaining:
                    last_part      = ordered[-1]["Part"]
                    last_color_val = part_color.get(last_part, "UNKNOWN")

                    def _co_cost(r, _lc=last_color_val):
                        p = r["Part"]
                        c = part_color.get(p, "UNKNOWN")
                        if p == last_part:
                            return (-1.0, p)
                        base  = vt_changeover.get(m, DEFAULT_CHANGEOVER_HRS)
                        purge = (
                            COLOR_PURGE_HRS
                            if _lc not in ("UNKNOWN", "NEEDS_PURGE")
                               and c not in ("UNKNOWN",)
                               and _lc != c
                            else 0.0
                        )
                        return (base + purge, p)

                    remaining.sort(key=_co_cost)
                    ordered.append(remaining.pop(0))
                for i, row in enumerate(ordered):
                    p         = row["Part"]
                    prev_part = yesterday_part if i == 0 else ordered[i - 1]["Part"]
                    if prev_part is None or prev_part == p:
                        new_co = 0.0
                    else:
                        base_co    = vt_changeover.get(m, DEFAULT_CHANGEOVER_HRS)
                        prev_color = ALL_KNOWN_COLORS.get(prev_part, "UNKNOWN")
                        new_color  = part_color.get(p, "UNKNOWN")
                        purge = (
                            COLOR_PURGE_HRS
                            if prev_color not in ("UNKNOWN", "NEEDS_PURGE")
                               and new_color not in ("UNKNOWN",)
                               and prev_color != new_color
                            else 0.0
                        )
                        new_co = base_co + purge
                    row["Changeover_Hrs"]  = round(new_co, 3)
                    row["Changeover"]      = "No" if new_co == 0 else "Yes"
                    row["Total_Hrs_Used"]  = round(new_co + float(row.get("Run_Hours", 0) or 0), 3)
                    row["Color_Purge"]     = "Yes" if (
                        new_co > vt_changeover.get(m, DEFAULT_CHANGEOVER_HRS) + 0.001
                    ) else "No"
                resequenced_plan.extend(ordered)
            resequenced_plan.extend(other_rows)
            return resequenced_plan

        # =========================================================
        # MACHINE HOURS RECONCILER
        # =========================================================

        def reconcile_machine_hours(plan, machine_hours):
            recomputed = {m: 0.0 for m in vt_machines}
            for row in plan:
                m = row.get("Machine")
                if m in recomputed:
                    recomputed[m] += float(row.get("Run_Hours", 0) or 0) + float(row.get("Changeover_Hrs", 0) or 0)
            for m in vt_machines:
                machine_hours[m] = round(recomputed[m], 4)

        # =========================================================
        # VALIDATION
        # =========================================================

        def validate_plan_rows(plan, current_inventory):
            violations = []
            for row in plan:
                p     = row["Part"]
                run_h = float(row.get("Run_Hours", 0) or 0)
                qty   = float(row.get("Production_Qty", 0) or 0)
                eff   = effective_daily.get(p, 0)
                v     = []
                if run_h < MIN_RUN_HOURS - 0.001:
                    v.append(f"Run_Hours={run_h:.3f} < MIN={MIN_RUN_HOURS}")
                if v:
                    row["Indent_Met"] = "VIOLATION: " + " | ".join(v)
                    violations.append({
                        "Part": p, "Machine": row.get("Machine", "—"),
                        "Run_Hours": run_h, "Qty": qty,
                        "Effective_Daily": eff, "Violations": " | ".join(v),
                    })
            return violations

        # =========================================================
        # FORWARD LOOK
        # =========================================================

        def compute_forward_look(current_inventory_after, all_parts, scenario_id):
            rows = []
            for p in sorted(all_parts):
                eff     = effective_daily.get(p, 0)
                monthly = indent_monthly.get(p, 0)
                if eff <= 0 or monthly <= 0:
                    continue
                inv_now = current_inventory_after.get(p, 0)
                days_now          = inv_now / eff if eff > 0 else 999
                days_until_safety = max(0.0, round((inv_now - SAFETY_DAYS * eff) / eff, 1))
                days_until_zero   = max(0.0, round(inv_now / eff, 1))
                alert = ""
                if days_until_zero <= FORWARD_LOOK_DAYS:
                    alert = f"ZERO-STOCK RISK in {days_until_zero:.1f} days"
                elif days_until_safety <= FORWARD_LOOK_DAYS:
                    alert = f"BELOW SAFETY in {days_until_safety:.1f} days"
                if alert:
                    rows.append({
                        "Part":           p,
                        "Color":          part_color.get(p, "UNKNOWN"),
                        "Category":       part_category.get(p, "Stranger"),
                        "Fixed_Machine":  part_fixed_machine.get(p, "—"),
                        "Indent_Daily":   round(indent_daily.get(p, 0), 2),
                        "Demand_Daily":   round(demand_daily.get(p, 0), 2),
                        "Effective_Daily": round(eff, 2),
                        "Inv_After_Today":     round(inv_now, 0),
                        "Days_Coverage_Today": round(days_now, 2),
                        "Days_Until_Safety":   days_until_safety,
                        "Days_Until_Zero":     days_until_zero,
                        "Alert":  alert,
                        "Action": "ESCALATE" if days_until_zero <= 2 else "Plan in next 1-3 days",
                    })
            df = pd.DataFrame(rows)
            if not df.empty:
                df = df.sort_values(["Days_Until_Zero", "Part"]).reset_index(drop=True)
            return df

        # =========================================================
        # OUTPUT VIEW BUILDERS
        # =========================================================

        def build_multi_machine_view(plan):
            if not plan:
                return pd.DataFrame()
            part_rows = defaultdict(list)
            for row in plan:
                part_rows[row["Part"]].append(row)
            multi = {p: rows for p, rows in part_rows.items() if len(rows) > 1}
            if not multi:
                return pd.DataFrame()
            output_rows = []
            for part, rows in sorted(multi.items(),
                                     key=lambda x: (-sum(r["Production_Qty"] for r in x[1]), x[0])):
                eff       = effective_daily.get(part, 0)
                total_qty = sum(float(r["Production_Qty"]) for r in rows)
                for row in sorted(rows, key=lambda r: r["Machine"]):
                    output_rows.append({
                        "Part":            part,
                        "Machines_Used":   len(rows),
                        "Machine":         row["Machine"],
                        "Role":            row.get("Role", "Primary"),
                        "Run_Hours":       round(float(row["Run_Hours"]), 2),
                        "Production_Qty":  round(float(row["Production_Qty"]), 0),
                        "Indent_Daily":    round(indent_daily.get(part, 0), 2),
                        "Demand_Daily":    round(demand_daily.get(part, 0), 2),
                        "Effective_Daily": round(eff, 2),
                        "Total_Qty_All_Machines": round(total_qty, 0),
                        "Type": row.get("Type", "—"),
                    })
            return pd.DataFrame(output_rows)

        def build_production_vs_indent(plan, all_parts):
            if not plan:
                return pd.DataFrame()
            part_qty      = defaultdict(float)
            part_machines = defaultdict(list)
            for row in plan:
                p = row["Part"]
                part_qty[p]      += float(row.get("Production_Qty", 0))
                part_machines[p].append(row["Machine"])
            rows = []
            for p in sorted(part_qty.keys()):
                i_daily  = indent_daily.get(p, 0)
                d_daily  = demand_daily.get(p, 0)
                eff      = effective_daily.get(p, 0)
                inv_b    = inventory.get(p, 0)
                produced = round(part_qty[p], 0)
                inv_after     = round(inv_b + produced, 0)
                gap_vs_indent = round(produced - i_daily, 0)
                gap_vs_demand = round(produced - d_daily, 0)
                gap_vs_eff    = round(produced - eff,     0)
                gap_dir = "OVER" if gap_vs_eff > 0 else ("UNDER" if gap_vs_eff < 0 else "MET")
                rows.append({
                    "Part":     p,
                    "Category": part_category.get(p, "Stranger"),
                    "Machines": ", ".join(sorted(dict.fromkeys(part_machines[p]))),
                    "Total_Qty_Produced":  produced,
                    "Indent_Daily":        round(i_daily, 2),
                    "Demand_Daily":        round(d_daily, 2),
                    "Effective_Daily":     round(eff,     2),
                    "Gap_vs_Indent":       gap_vs_indent,
                    "Gap_vs_Demand":       gap_vs_demand,
                    "Gap_vs_Effective":    gap_vs_eff,
                    "Gap_Direction":       gap_dir,
                    "Inventory_Before":    round(inv_b,   0),
                    "Inventory_After":     inv_after,
                    "Days_Coverage_After": round(inv_after / eff, 2) if eff > 0 else 0,
                })
            df = pd.DataFrame(rows)
            if not df.empty:
                order_map = {"UNDER": 0, "MET": 1, "OVER": 2}
                df["_sort"] = df["Gap_Direction"].map(order_map)
                df = df.sort_values(["_sort", "Gap_vs_Effective", "Part"]).drop(
                    columns=["_sort"]).reset_index(drop=True)
            return df

        def build_inventory_target_sheet(plan, all_parts, scenario_id):
            part_produced = defaultdict(float)
            for row in plan:
                part_produced[row["Part"]] += float(row.get("Production_Qty", 0))
            rows = []
            for p in sorted(all_parts):
                i_daily  = indent_daily.get(p, 0)
                d_daily  = demand_daily.get(p, 0)
                eff      = effective_daily.get(p, 0)
                inv_b    = inventory.get(p, 0)
                produced  = round(part_produced.get(p, 0), 0)
                inv_after = round(inv_b + produced, 0)
                days_after = round(inv_after / eff, 2) if eff > 0 else 0
                target_qty = round(TARGET_DAYS * eff, 0)
                if inv_after == 0:
                    status = "CRITICAL"
                elif days_after < SAFETY_DAYS:
                    status = "BELOW_SAFETY"
                elif days_after < TARGET_DAYS:
                    status = "BUILDING"
                else:
                    status = "AT_TARGET"
                rows.append({
                    "Part":           p,
                    "Category":       part_category.get(p, "Stranger"),
                    "Indent_Daily":   round(i_daily, 2),
                    "Demand_Daily":   round(d_daily, 2),
                    "Effective_Daily": round(eff,    2),
                    "Inv_Before":     round(inv_b,   0),
                    "Produced_Today": produced,
                    "Inv_After":      inv_after,
                    "Days_Coverage_After": days_after,
                    "Target_Qty_5days":    target_qty,
                    "Buffer_Status":       status,
                    "Scheduled_Today": "YES" if produced > 0 else "NO",
                })
            df = pd.DataFrame(rows)
            if not df.empty:
                status_order = {"CRITICAL": 0, "BELOW_SAFETY": 1, "BUILDING": 2, "AT_TARGET": 3}
                df["_sort"] = df["Buffer_Status"].map(status_order)
                df = df.sort_values(["_sort", "Part"], ascending=True).drop(
                    columns=["_sort"]).reset_index(drop=True)
            return df

        def compute_indent_horizon(parts):
            rows = []
            for p in sorted(parts):
                inv     = inventory.get(p, 0.0)
                monthly = indent_monthly.get(p, 0.0)
                i_daily = indent_daily.get(p, 0.0)
                d_daily = demand_daily.get(p, 0.0)
                eff     = effective_daily.get(p, 0.0)
                r       = rate.get(p, 1.0)
                skip, skip_reason = should_skip(p)
                days_cov = inv / eff if eff > 0 else 0
                if inv == 0.0 and monthly > 0:
                    status = "ZERO INV"
                elif skip and inv >= TARGET_DAYS * eff:
                    status = "AT TARGET"
                elif skip:
                    status = "SKIPPED"
                else:
                    status = "PRODUCTION NEEDED"
                rows.append({
                    "Part":            p,
                    "Monthly_Indent":  round(monthly, 0),
                    "Indent_Daily":    round(i_daily, 2),
                    "Demand_Daily":    round(d_daily, 2),
                    "Effective_Daily": round(eff,     2),
                    "Inventory_Now":   round(inv,     0),
                    "Days_Coverage":   round(days_cov, 2),
                    "Indent_Status":   status,
                    "Skip_Reason":     skip_reason,
                })
            return pd.DataFrame(rows)

        def build_terminal_status_sheet():
            rows = []
            all_terminals_known = set(terminal_status.keys())
            for terms in part_terminals.values():
                for t in terms:
                    all_terminals_known.add(t)
            for t in sorted(all_terminals_known):
                inv = terminal_status.get(t, None)
                parts_needing = sorted([p for p, terms in part_terminals.items() if t in terms])
                parts_blocked = []
                for p in parts_needing:
                    cat    = part_category.get(p, "Stranger")
                    r_val  = rate.get(p, 0.0)
                    pct    = TERMINAL_THRESHOLD.get(cat, 0.75)
                    thresh = pct * MIN_RUN_HOURS * r_val
                    if inv is None or inv < thresh:
                        parts_blocked.append(
                            f"{p}(need>={thresh:.0f} [{cat} {int(pct*100)}%×{MIN_RUN_HOURS}h×{r_val:.1f}/h])"
                        )
                rows.append({
                    "Terminal":        t,
                    "Inventory":       round(inv, 0) if inv is not None else "NOT IN SHEET",
                    "Parts_Requiring": ", ".join(parts_needing) if parts_needing else "—",
                    "Parts_Blocked":   ", ".join(parts_blocked) if parts_blocked else "—",
                    "Impact": "BLOCKING" if parts_blocked else ("Adequate" if parts_needing else "No parts"),
                })
            return pd.DataFrame(rows)

        def build_fixed_machine_status(plan, current_inventory, phase_a_machines):
            rows = []
            for machine in sorted(machine_fixed_parts.keys()):
                fixed_parts = machine_fixed_parts[machine]
                hrs_used = sum(
                    float(r.get("Run_Hours", 0)) + float(r.get("Changeover_Hrs", 0))
                    for r in plan if r["Machine"] == machine
                )
                for p in sorted(fixed_parts):
                    i_daily  = indent_daily.get(p, 0)
                    d_daily  = demand_daily.get(p, 0)
                    eff      = effective_daily.get(p, 0)
                    inv_b    = inventory.get(p, 0)
                    inv_now  = current_inventory.get(p, inv_b)
                    days_now = round(inv_now / eff, 2) if eff > 0 else 0
                    produced = round(inv_now - inv_b, 0)
                    rows.append({
                        "Machine":       machine,
                        "Phase":         "Lockdown (always runs)",
                        "Part":          p,
                        "Indent_Daily":  round(i_daily, 2),
                        "Demand_Daily":  round(d_daily, 2),
                        "Effective_Daily": round(eff,   2),
                        "Inv_Before":    round(inv_b,   0),
                        "Produced_Today": produced,
                        "Inv_After":     round(inv_now, 0),
                        "Days_Coverage_After": days_now,
                        "Machine_Hrs_Used": round(hrs_used, 2),
                    })
            return pd.DataFrame(rows)

        def build_machine_wise_plan(plan_df):
            if plan_df.empty:
                return pd.DataFrame()
            rows = []
            for m in sorted(vt_machines):
                machine_rows = plan_df[plan_df["Machine"] == m].copy()
                if machine_rows.empty:
                    continue
                for _, pr in machine_rows.iterrows():
                    p = pr.get("Part", "—")
                    rows.append({
                        "Machine":    m,
                        "Part":       p,
                        "Color":      part_color.get(p, "UNKNOWN"),
                        "Category":   part_category.get(p, "Stranger"),
                        "Role":       pr.get("Role", "Primary"),
                        "Run_Hours":  round(float(pr.get("Run_Hours", 0) or 0), 2),
                        "Changeover_Hrs": round(float(pr.get("Changeover_Hrs", 0) or 0), 3),
                        "Production_Qty": round(float(pr.get("Production_Qty", 0) or 0), 0),
                        "Indent_Daily":   round(float(pr.get("Indent_Daily", 0) or 0), 2),
                        "Demand_Daily":   round(float(pr.get("Demand_Daily", 0) or 0), 2),
                        "Effective_Daily": round(float(pr.get("Effective_Daily", 0) or 0), 2),
                        "Type":       pr.get("Type", "Primary") or "Primary",
                        "Row_Type":   "Part",
                    })
                co_total  = machine_rows["Changeover_Hrs"].apply(lambda x: float(x) if pd.notna(x) else 0).sum()
                run_total = machine_rows["Run_Hours"].apply(lambda x: float(x) if pd.notna(x) else 0).sum()
                qty_total = machine_rows["Production_Qty"].apply(lambda x: float(x) if pd.notna(x) else 0).sum()
                hrs_total = round(co_total + run_total, 2)
                rows.append({
                    "Machine": m, "Part": f"TOTAL — {m}",
                    "Color": "—", "Category": "—", "Role": "—",
                    "Run_Hours":     round(run_total, 2),
                    "Changeover_Hrs": round(co_total, 2),
                    "Production_Qty": round(qty_total, 0),
                    "Indent_Daily": "—", "Demand_Daily": "—", "Effective_Daily": "—",
                    "Type": f"Total {hrs_total}h / {AVAILABLE_HOURS}h  |  Util {round(hrs_total/AVAILABLE_HOURS*100,1)}%",
                    "Row_Type": "Summary",
                })
                rows.append({k: "" for k in rows[-1].keys()})
            return pd.DataFrame(rows)

        def build_co_queue(plan, machines):
            events = _collect_co_events(plan, sorted(machines))
            if not events:
                return pd.DataFrame()
            events.sort(key=lambda e: (e["natural_start"], e["machine"]))
            rows = []
            tool_changer_free_at = 0.0
            for pos, ev in enumerate(events, 1):
                natural_start = _recompute_natural_start(ev, plan)
                co_h          = ev["co_duration"]
                actual_start  = max(natural_start, tool_changer_free_at)
                tool_changer_free_at = actual_start + co_h
                before_color  = part_color.get(ev["part_before"], "UNKNOWN")
                after_color   = part_color.get(ev["part_after"],  "UNKNOWN")
                color_change  = (before_color != after_color
                                 and before_color != "UNKNOWN"
                                 and after_color  != "UNKNOWN")
                rows.append({
                    "Queue_Position":  pos,
                    "Machine":         ev["machine"],
                    "Part_Before":     ev["part_before"],
                    "Part_After":      ev["part_after"],
                    "Color_Before":    before_color,
                    "Color_After":     after_color,
                    "Color_Change":    "YES — PURGE" if color_change else "No",
                    "CO_Duration_Min": round(co_h * 60, 1),
                })
            return pd.DataFrame(rows)

        # =========================================================
        # DEMAND COMPLIANCE SHEET
        # =========================================================

        def build_demand_compliance(plan, all_parts):
            part_produced = defaultdict(float)
            for row in plan:
                part_produced[row["Part"]] += float(row.get("Production_Qty", 0))

            rows = []
            for p in sorted(all_parts):
                i_daily  = indent_daily.get(p, 0)
                d_daily  = demand_daily.get(p, 0)
                eff      = effective_daily.get(p, 0)
                inv_b    = inventory.get(p, 0)
                produced = round(part_produced.get(p, 0), 0)
                inv_after = inv_b + produced

                inv_covers_demand   = inv_b >= d_daily
                total_covers_demand = inv_after >= d_daily

                demand_gap = max(0.0, d_daily - inv_after)
                indent_gap = max(0.0, i_daily - produced)

                if d_daily == 0:
                    compliance = "NO DEMAND"
                elif total_covers_demand:
                    compliance = "DEMAND MET"
                elif inv_covers_demand:
                    compliance = "MET BY INVENTORY"
                else:
                    compliance = "DEMAND GAP"

                rows.append({
                    "Part":            p,
                    "Category":        part_category.get(p, "Stranger"),
                    "Indent_Daily":    round(i_daily,  2),
                    "Demand_Daily":    round(d_daily,  2),
                    "Effective_Daily": round(eff,      2),
                    "Inventory_SOD":   round(inv_b,    0),
                    "Produced_Today":  produced,
                    "Inv_Available":   round(inv_after, 0),
                    "Demand_Gap":      round(demand_gap, 0),
                    "Indent_Gap":      round(indent_gap, 0),
                    "Inv_Covers_Demand": "YES" if inv_covers_demand else "NO",
                    "Compliance":      compliance,
                })

            df = pd.DataFrame(rows)
            if not df.empty:
                order = {"DEMAND GAP": 0, "DEMAND MET": 1, "MET BY INVENTORY": 2, "NO DEMAND": 3}
                df["_sort"] = df["Compliance"].map(order)
                df = df.sort_values(["_sort", "Demand_Gap", "Part"], ascending=[True, False, True])
                df = df.drop(columns=["_sort"]).reset_index(drop=True)
            return df

        # =========================================================
        # PART AUDIT
        # =========================================================

        def build_part_audit(data, data_zero_rate, vt_matrix, matrix_parts=None):
            if matrix_parts is None:
                matrix_parts = set(str(p).strip() for p in vt_matrix["Part"] if pd.notna(p))
            zero_rate_set = set(data_zero_rate["Material"].unique())
            audit_rows    = []
            for part in sorted(data["Material"].unique()):
                inv     = inventory.get(part, 0.0)
                r_val   = rate.get(part, None)
                monthly = indent_monthly.get(part, 0.0)
                i_daily = indent_daily.get(part, 0.0)
                d_daily = demand_daily.get(part, 0.0)
                eff     = effective_daily.get(part, 0.0)
                days_cov = inv / eff if eff > 0 else 0
                t_blk, t_rsn = terminal_blocked(part)
                if part in zero_rate_set or r_val is None:
                    status = "ZERO/MISSING CYCLE TIME"
                elif part not in matrix_parts:
                    status = f"NOT IN {S}_MATRIX"
                elif monthly == 0:
                    status = "ZERO INDENT"
                elif i_daily <= MIN_DAILY_INDENT and d_daily <= MIN_DAILY_INDENT:
                    status = "SKIPPED (LOW INDENT+DEMAND)"
                elif eff > 0 and inv >= TARGET_DAYS * eff:
                    status = "AT TARGET — SKIP"
                elif t_blk:
                    status = "BLOCKED — TERMINAL"
                else:
                    status = "ENTERS SCHEDULER"
                audit_rows.append({
                    "Part":            part,
                    "Color":           part_color.get(part, "UNKNOWN"),
                    "Monthly_Indent":  round(monthly, 0),
                    "Indent_Daily":    round(i_daily, 2),
                    "Demand_Daily":    round(d_daily, 2),
                    "Effective_Daily": round(eff,     2),
                    "Inventory":       round(inv,     0),
                    "Days_Coverage":   round(days_cov, 2),
                    "Rate_Per_Hour":   round(r_val, 2) if r_val else "—",
                    "Status":          status,
                })
            return pd.DataFrame(audit_rows)

        # =========================================================
        # MAIN SCHEDULER
        # =========================================================

        def schedule(parts, label=""):
            scenario_id, scenario_desc = classify_scenario(parts)
            horizon_df   = compute_indent_horizon(parts)
            active_parts = sorted([
                p for p in parts if not is_scheduling_skip(p) and indent_monthly.get(p, 0) > 0
            ])
            priority_scores, score_rows = compute_priority_scores(active_parts)
            score_df = pd.DataFrame(score_rows) if score_rows else pd.DataFrame()

            machine_hours     = {m: 0.0 for m in sorted(vt_machines)}
            machine_last_part = {m: machine_state.get(m) for m in sorted(vt_machines)}
            current_inventory = dict(sorted(inventory.items()))
            inventory_start_of_day = dict(sorted(inventory.items()))

            plan            = []
            already_planned = set()
            not_planned     = []
            deferred        = []

            # V13: Fixed machine pass — absolute lockdown, returns all fixed machines as phase_a
            phase_a_machines, _ = schedule_fixed_machines(
                machine_hours, machine_last_part, current_inventory,
                plan, already_planned, priority_scores, scenario_id)

            zero_inv_rr = sorted([
                p for p in active_parts
                if current_inventory.get(p, 0) == 0 and p not in already_planned
                and part_category.get(p, "Stranger") in ("Runner", "Repeater")
                and vt_compat.get(p)
            ])

            # V13: sort active parts by priority score (high demand × low coverage first)
            sorted_active = sorted(
                [p for p in active_parts if p not in already_planned],
                key=lambda p: (-priority_scores.get(p, 0), -effective_daily.get(p, 0), p),
            )

            for part in sorted_active:
                monthly = indent_monthly.get(part, 0)
                if monthly == 0:
                    deferred.append({"Part": part, "Reason": "Monthly indent = 0"})
                    continue
                t_blocked, t_reason = terminal_blocked(part)
                if t_blocked:
                    not_planned.append({"Part": part, "Reason": t_reason})
                    continue
                if not vt_compat.get(part):
                    not_planned.append({"Part": part, "Reason": f"Not in {S}_Matrix"})
                    continue
                new_rows = assign_part(part, scenario_id, machine_hours, machine_last_part,
                                       current_inventory, plan, already_planned, priority_scores)
                if new_rows:
                    plan.extend(new_rows)
                else:
                    not_planned.append({"Part": part, "Reason": "No capacity"})

            unscheduled_zero = [p for p in zero_inv_rr if p not in already_planned]
            for part in unscheduled_zero:
                success = displace_for_zero_inv(
                    part, machine_hours, machine_last_part,
                    current_inventory, plan, already_planned, priority_scores, phase_a_machines)
                if success:
                    not_planned[:] = [r for r in not_planned if r.get("Part") != part]

            runner_priority_log = enforce_runner_priority(
                plan, machine_hours, machine_last_part, current_inventory,
                already_planned, priority_scores, inventory_start_of_day)
            newly_planned = {r["Runner_Part"] for r in runner_priority_log if "PLANNED" in r.get("Result", "")}
            not_planned[:] = [r for r in not_planned if r.get("Part") not in newly_planned]

            micro_idle = utilization_enforcer(
                plan, machine_hours, machine_last_part, list(parts),
                already_planned, current_inventory, scenario_id, priority_scores, phase_a_machines)

            strategic_buffer_filler(plan, machine_hours, machine_last_part,
                                    list(parts), already_planned, current_inventory,
                                    scenario_id, phase_a_machines)

            reconcile_machine_hours(plan, machine_hours)
            plan = resequence_machine_rows(plan, machine_state)
            reconcile_machine_hours(plan, machine_hours)

            for m in sorted(vt_machines):
                m_rows = [r for r in plan if r["Machine"] == m]
                if m_rows:
                    machine_last_part[m] = m_rows[-1]["Part"]

            stagger_changeovers(plan, vt_machines, machine_hours)
            reconcile_machine_hours(plan, machine_hours)

            violations           = validate_plan_rows(plan, current_inventory)
            forward_look_df      = compute_forward_look(current_inventory, list(parts), scenario_id)
            multi_machine_df     = build_multi_machine_view(plan)
            prod_vs_indent_df    = build_production_vs_indent(plan, list(parts))
            inv_target_df        = build_inventory_target_sheet(plan, list(parts), scenario_id)
            demand_compliance_df = build_demand_compliance(plan, list(parts))
            runner_priority_df   = pd.DataFrame(runner_priority_log) if runner_priority_log else pd.DataFrame()
            fixed_machine_status_df = build_fixed_machine_status(plan, current_inventory, phase_a_machines)
            violations_df        = pd.DataFrame(violations) if violations else pd.DataFrame()

            mach_rows = []
            for m in sorted(vt_machines):
                used       = machine_hours.get(m, 0)
                parts_run  = sorted({r["Part"] for r in plan if r["Machine"] == m})
                co_count   = sum(1 for r in plan if r["Machine"] == m and r.get("Changeover") == "Yes")
                util_pct   = round(used / AVAILABLE_HOURS * 100, 1)
                is_fixed_m = m in machine_fixed_parts
                mach_rows.append({
                    "Machine":      m,
                    "Fixed_Machine": "YES" if is_fixed_m else "No",
                    "Used_Hours":   round(used, 2),
                    "Unused_Hours": round(AVAILABLE_HOURS - used, 2),
                    "Utilization_%": util_pct,
                    "CO_Count":     co_count,
                    "Status": (
                        "FULL"     if used >= AVAILABLE_HOURS - 0.3 else
                        "GOOD"     if util_pct >= 98 else
                        "OK"       if util_pct >= 90 else "UNDERUSED"
                    ),
                    "Parts_Planned": len(parts_run),
                    "Last_Part":    machine_last_part.get(m) or "—",
                    "All_Parts":    ", ".join(parts_run) if parts_run else "— idle —",
                })

            inv_rows = []
            for p in sorted(parts):
                i_daily  = indent_daily.get(p, 0)
                d_daily  = demand_daily.get(p, 0)
                eff      = effective_daily.get(p, 0)
                inv_b    = inventory.get(p, 0)
                produced = sum(float(r["Production_Qty"]) for r in plan if r["Part"] == p)
                inv_after = inv_b + produced
                days_cov  = inv_after / eff if eff > 0 else 0
                inv_rows.append({
                    "Part":            p,
                    "Indent_Daily":    round(i_daily,  2),
                    "Demand_Daily":    round(d_daily,  2),
                    "Effective_Daily": round(eff,      2),
                    "Inv_Before":      round(inv_b,    0),
                    "Produced_Today":  round(produced, 0),
                    "Inv_After":       round(inv_after, 0),
                    "Days_Coverage":   round(days_cov,  2),
                    "Status": (
                        "AT_TARGET" if days_cov >= TARGET_DAYS else
                        "OK"        if days_cov >= SAFETY_DAYS else
                        "LOW"       if days_cov >= 1 else "CRITICAL"
                    ),
                })

            micro_df  = pd.DataFrame(micro_idle) if micro_idle else pd.DataFrame()
            plan_df   = pd.DataFrame(plan) if plan else pd.DataFrame()
            def_df    = pd.DataFrame(deferred)    if deferred    else pd.DataFrame()
            not_df    = pd.DataFrame(not_planned) if not_planned else pd.DataFrame()
            mach_df   = pd.DataFrame(mach_rows)
            inv_df    = pd.DataFrame(inv_rows)

            if not plan_df.empty:
                plan_df.insert(0, "Planning_Date", str(planning_date))
                plan_df.insert(1, "Scenario", scenario_desc)

            return (plan_df, def_df, not_df, mach_df, inv_df, machine_last_part,
                    horizon_df, pd.DataFrame(), score_df, micro_df, multi_machine_df,
                    prod_vs_indent_df, inv_target_df, runner_priority_df,
                    fixed_machine_status_df, forward_look_df, violations_df,
                    phase_a_machines, demand_compliance_df)

        # =========================================================
        # RUN
        # =========================================================

        matrix_parts_set = set(str(p).strip() for p in vt_matrix["Part"] if pd.notna(p))
        vt_parts = data_valid[data_valid["Material"].isin(matrix_parts_set)]["Material"].unique()

        (vt_plan, vt_def, vt_not, vt_mach, vt_inv,
         vt_state, vt_horizon, _unused,
         vt_scores, vt_micro, vt_multi_machine,
         vt_prod_vs_indent, vt_inv_target,
         vt_runner_priority, vt_fixed_status,
         vt_forward_look, vt_violations,
         phase_a_machines,
         vt_demand_compliance) = schedule(vt_parts, f"{S} Machines")

        save_machine_state(vt_state)

        vt_mw   = build_machine_wise_plan(vt_plan)
        vt_co_q = build_co_queue(
            [{k: v for k, v in r.items()} for r in vt_plan.to_dict("records")] if not vt_plan.empty else [],
            vt_machines,
        )
        audit_df              = build_part_audit(data, data_zero_rate, vt_matrix, matrix_parts_set)
        vt_terminal_status_df = build_terminal_status_sheet()

        # =========================================================
        # WRITE EXCEL
        # =========================================================

        write_excel(
            output_path=output_path,
            vt_plan=vt_plan, vt_mw=vt_mw, vt_co_q=vt_co_q,
            vt_def=vt_def, vt_not=vt_not, vt_mach=vt_mach, vt_inv=vt_inv,
            vt_horizon=vt_horizon, vt_scores=vt_scores, vt_micro=vt_micro,
            vt_multi_machine=vt_multi_machine, vt_prod_vs_indent=vt_prod_vs_indent,
            vt_inv_target=vt_inv_target, vt_runner_priority=vt_runner_priority,
            vt_fixed_status=vt_fixed_status, vt_forward_look=vt_forward_look,
            vt_violations=vt_violations, audit_df=audit_df,
            vt_terminal_status_df=vt_terminal_status_df,
            vt_demand_compliance=vt_demand_compliance,
            available_hours=AVAILABLE_HOURS,
            section=S,
        )

        # =========================================================
        # BUILD RETURN VALUE
        # =========================================================

        def _df_to_list(df):
            if df is None or df.empty:
                return []
            return df.where(pd.notnull(df), None).to_dict("records")

        summary = {
            "section":               S,
            "planning_date":         str(planning_date),
            "indent_month":          str(indent_month),
            "working_days":          WORKING_DAYS,
            "total_parts_in_sheet":  len(data),
            "parts_with_valid_rate": len(data_valid),
            "demand_file_loaded":    bool(demand_raw_daily),
            "parts_with_demand":     sum(1 for p in effective_daily if demand_raw_daily.get(p, 0) > indent_daily.get(p, 0)),
            "plan_rows":             len(vt_plan),
            "not_planned":           len(vt_not),
            "deferred":              len(vt_def),
            "violations":            len(vt_violations),
            "demand_gaps":           int((vt_demand_compliance["Compliance"] == "DEMAND GAP").sum())
                                     if not vt_demand_compliance.empty else 0,
            "avg_utilization_pct": (
                round(vt_mach["Utilization_%"].mean(), 1) if not vt_mach.empty else None
            ),
            "fixed_machines_locked": list(sorted(machine_fixed_parts.keys())),
            "terminal_thresholds":   {k: f"{int(v*100)}% × MIN_RUN_HOURS × rate" for k, v in TERMINAL_THRESHOLD.items()},
        }

        sheets = {
            f"{S}_Plan_By_Machine":      _df_to_list(vt_mw),
            f"{S}_CO_Queue":             _df_to_list(vt_co_q),
            f"{S}_Plan":                 _df_to_list(vt_plan),
            f"{S}_Demand_Compliance":    _df_to_list(vt_demand_compliance),
            f"{S}_Fixed_Machine_Status": _df_to_list(vt_fixed_status),
            f"{S}_Runner_Priority_Log":  _df_to_list(vt_runner_priority),
            f"{S}_Inventory_Target":     _df_to_list(vt_inv_target),
            f"{S}_Forward_Look":         _df_to_list(vt_forward_look),
            f"{S}_Multi_Machine_Parts":  _df_to_list(vt_multi_machine),
            f"{S}_Production_vs_Indent": _df_to_list(vt_prod_vs_indent),
            f"{S}_Priority_Scores":      _df_to_list(vt_scores),
            f"{S}_Machine_Util":         _df_to_list(vt_mach),
            f"{S}_Not_Planned":          _df_to_list(vt_not),
            f"{S}_Deferred":             _df_to_list(vt_def),
            f"{S}_Inventory_Health":     _df_to_list(vt_inv),
            f"{S}_Indent_Horizon":       _df_to_list(vt_horizon),
            f"{S}_Part_Audit":           _df_to_list(audit_df),
            f"{S}_Terminal_Status":      _df_to_list(vt_terminal_status_df),
            f"{S}_Micro_Idle":           _df_to_list(vt_micro),
            f"{S}_Violations":           _df_to_list(vt_violations),
        }

        return {
            "success":     True,
            "error":       None,
            "output_path": output_path,
            "summary":     summary,
            "sheets":      sheets,
        }

    except Exception as e:
        import traceback
        return {
            "success":   False,
            "error":     str(e),
            "traceback": traceback.format_exc(),
            "output_path": None,
            "summary": {},
            "sheets":  {},
        }


# =============================================================
# EXCEL WRITER
# =============================================================

def write_excel(
    output_path,
    vt_plan, vt_mw, vt_co_q, vt_def, vt_not, vt_mach, vt_inv,
    vt_horizon, vt_scores, vt_micro, vt_multi_machine,
    vt_prod_vs_indent, vt_inv_target, vt_runner_priority,
    vt_fixed_status, vt_forward_look, vt_violations,
    audit_df, vt_terminal_status_df,
    vt_demand_compliance=None,
    available_hours=22,
    section="VT",
):
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter

    S = (str(section).strip() if section is not None else "") or "VT"

    HEADER_COLORS = {
        f"{S}_Plan_By_Machine":    "0D6E6E",
        f"{S}_CO_Queue":           "375623",
        f"{S}_Plan":               "1F4E79",
        f"{S}_Demand_Compliance":  "7C3AED",
        f"{S}_Multi_Machine_Parts":"4A235A",
        f"{S}_Production_vs_Indent":"154360",
        f"{S}_Inventory_Target":   "1B4F72",
        f"{S}_Priority_Scores":    "2C4770",
        f"{S}_Machine_Util":       "375623",
        f"{S}_Not_Planned":        "7B2C2C",
        f"{S}_Deferred":           "7F6000",
        f"{S}_Inventory_Health":   "4A235A",
        f"{S}_Indent_Horizon":     "154360",
        f"{S}_Part_Audit":         "1C3557",
        f"{S}_Micro_Idle":         "5C3D2E",
        f"{S}_Terminal_Status":    "7B1C1C",
        f"{S}_Runner_Priority_Log":"7B3F00",
        f"{S}_Fixed_Machine_Status":"1A5276",
        f"{S}_Forward_Look":       "6D28D9",
        f"{S}_Violations":         "991B1B",
    }

    STATUS_FILLS = {
        "FULL":            PatternFill("solid", fgColor="C6EFCE"),
        "GOOD":            PatternFill("solid", fgColor="DDEBF7"),
        "OK":              PatternFill("solid", fgColor="EBF5E1"),
        "UNDERUSED":       PatternFill("solid", fgColor="FFC7CE"),
        "AT_TARGET":       PatternFill("solid", fgColor="C6EFCE"),
        "BUILDING":        PatternFill("solid", fgColor="DDEBF7"),
        "BELOW_SAFETY":    PatternFill("solid", fgColor="FFEB9C"),
        "CRITICAL":        PatternFill("solid", fgColor="FFC7CE"),
        "LOW":             PatternFill("solid", fgColor="FFEB9C"),
        "OVER":            PatternFill("solid", fgColor="DDEBF7"),
        "UNDER":           PatternFill("solid", fgColor="FFC7CE"),
        "MET":             PatternFill("solid", fgColor="C6EFCE"),
        "YES — PURGE":     PatternFill("solid", fgColor="FFC7CE"),
        "BLOCKING":        PatternFill("solid", fgColor="FFC7CE"),
        "Adequate":        PatternFill("solid", fgColor="C6EFCE"),
        "ENTERS SCHEDULER":PatternFill("solid", fgColor="C6EFCE"),
        "DEMAND MET":      PatternFill("solid", fgColor="C6EFCE"),
        "MET BY INVENTORY":PatternFill("solid", fgColor="EBF5E1"),
        "DEMAND GAP":      PatternFill("solid", fgColor="FFC7CE"),
        "NO DEMAND":       PatternFill("solid", fgColor="F2F2F2"),
    }

    def style_sheet(ws, header_hex):
        for cell in ws[1]:
            cell.fill = PatternFill("solid", fgColor=header_hex)
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[1].height = 32
        for col in ws.columns:
            col_letter = get_column_letter(col[0].column)
            max_len = max((len(str(c.value)) for c in col if c.value is not None), default=8)
            ws.column_dimensions[col_letter].width = max(10, min(55, max_len + 3))
        headers     = [cell.value for cell in ws[1]]
        status_cols = ["Status", "Gap_Direction", "Buffer_Status", "Color_Change",
                       "Impact", "Result", "Compliance"]
        for col_idx, col_name in enumerate(headers, start=1):
            if col_name and any(x in str(col_name) for x in status_cols):
                for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                    for cell in row:
                        fill = STATUS_FILLS.get(str(cell.value))
                        if fill:
                            cell.fill = fill
        ws.freeze_panes = "A2"

    def style_machine_wise_sheet(ws, available_hours):
        header_fill  = PatternFill("solid", fgColor="1F4E79")
        summary_fill = PatternFill("solid", fgColor="0D9488")
        part_fills   = [PatternFill("solid", fgColor="EFF6FF"), PatternFill("solid", fgColor="F0FDF4")]
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[1].height = 30
        headers          = [cell.value for cell in ws[1]]
        row_type_col     = headers.index("Row_Type") + 1 if "Row_Type" in headers else None
        machine_col      = headers.index("Machine")  + 1 if "Machine"  in headers else None
        machine_color_idx = 0
        current_machine   = None
        for row in ws.iter_rows(min_row=2):
            row_type = row[row_type_col - 1].value if row_type_col else ""
            machine  = row[machine_col  - 1].value if machine_col  else ""
            if machine and machine != current_machine:
                current_machine   = machine
                machine_color_idx = (machine_color_idx + 1) % 2
            if row_type == "Summary":
                for cell in row:
                    cell.fill = summary_fill
                    cell.font = Font(bold=True, color="FFFFFF", size=11)
            elif row_type == "Part":
                for cell in row:
                    cell.fill = part_fills[machine_color_idx]
        for col in ws.columns:
            col_letter = get_column_letter(col[0].column)
            max_len = max((len(str(c.value)) for c in col if c.value), default=8)
            ws.column_dimensions[col_letter].width = max(10, min(45, max_len + 3))
        ws.freeze_panes = "B2"

    plan_by_machine_sheet = f"{S}_Plan_By_Machine"
    sheets = {
        plan_by_machine_sheet:         vt_mw,
        f"{S}_CO_Queue":               vt_co_q,
        f"{S}_Plan":                   vt_plan,
        f"{S}_Demand_Compliance":      vt_demand_compliance,
        f"{S}_Fixed_Machine_Status":   vt_fixed_status,
        f"{S}_Runner_Priority_Log":    vt_runner_priority,
        f"{S}_Inventory_Target":       vt_inv_target,
        f"{S}_Forward_Look":           vt_forward_look,
        f"{S}_Multi_Machine_Parts":    vt_multi_machine,
        f"{S}_Production_vs_Indent":   vt_prod_vs_indent,
        f"{S}_Priority_Scores":        vt_scores,
        f"{S}_Machine_Util":           vt_mach,
        f"{S}_Not_Planned":            vt_not,
        f"{S}_Deferred":               vt_def,
        f"{S}_Inventory_Health":       vt_inv,
        f"{S}_Indent_Horizon":         vt_horizon,
        f"{S}_Part_Audit":             audit_df,
        f"{S}_Terminal_Status":        vt_terminal_status_df,
    }
    if vt_micro is not None and not vt_micro.empty:
        sheets[f"{S}_Micro_Idle"] = vt_micro
    if vt_violations is not None and not vt_violations.empty:
        sheets[f"{S}_Violations"] = vt_violations

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for sheet_name, df in sheets.items():
            if df is not None and not df.empty:
                df.to_excel(writer, sheet_name=sheet_name, index=False)

    wb = load_workbook(output_path)

    if plan_by_machine_sheet in wb.sheetnames:
        style_machine_wise_sheet(wb[plan_by_machine_sheet], available_hours)

    for sheet_name, header_hex in HEADER_COLORS.items():
        if sheet_name in wb.sheetnames and sheet_name != plan_by_machine_sheet:
            style_sheet(wb[sheet_name], header_hex)

    for name, color in HEADER_COLORS.items():
        if name in wb.sheetnames:
            wb[name].sheet_properties.tabColor = color

    wb.save(output_path)